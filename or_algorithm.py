import pandas as pd
import datetime
import pymysql
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from ortools.linear_solver import pywraplp
from data_processing import DataProcessing
from app import app, db
from sqlalchemy import text
from models import Timetable, User

"""
To-Do Liste:

Prio 1:

 - (erl.) NB6: Max. einen Arbeitsblock pro Tag
 - (erl.) die calc_time soll automatisch errechnet werden
 - (erl.) Als Key oder i für MA soll nicht mehr MA1, MA2 usw stehen, sondern die user_id (zB. 1002)
 - (erl.) Shifts/Employment_level aus der Datenbank ziehen
 - (erl.) auf Viertelstunden wechseln
 - (erl.) Gerechte Verteilung Anpassen, das Stunden von "Perm" Mitarbeiter abgezogen werden
 - (erl.) Weiche NB4 implementieren, hat noch nicht wunschgemäss geklappt
 - (erl.) Den Übergang auf harte und weiche NBs machen? 
 - (erl.) Die gesolvten Daten in der Datenbank speichern
 - (erl.) Daten für Solven in die Datenbank einpflegen (max. Zeit, min. Zeit, Solvingzeitraum, Toleranz für die Stundenverteilung, ...)
 - (erl.) Eine if Anweseiung, wenn der Betrieb an einem Tag geschlossen hat. Dann soll an diesem Tag nicht gesolvet werden


 To-Do's 
 -------------------------------
 - NB9 mit 3 Schichten fertigbauen
 - (erl.) Stunden Teiler für 1/4, 1/2 und 1h einbauen
 - Opening Hour 2 einbauen
 - Code ordnen
 - Testphase

 - Bei Company Shifts darf es nur die Möglichkeit geben, 1,2 oder 3 Schichten anzuwählen
 - "Gewünschte max. Zeit pro Woche" in Solver Req muss gelöscht werden
 - Der erstellte "divisor" in data_processing könnte als Attribut initialisiert werden, damit es nicht bei jeder Methode einzeln berechnet werden muss

 
 Fragen an die Runde:
 -------------------------------
 - MA mit verschiedenen Profilen - Department (Koch, Service, ..)? Wie genau lösen wir das?
 - Was machen wenn nicht genug Stunden von MA eingegeben wurden? Einen virtuellen MA der anzeigt, an welchen Stunden noch MA eingeteilt werden müssen?
   Andere Vorschläge?
 - Sobald gesolvt wird, kann die Webseite nicht mehr bedient werden. Wie können wir das lösen?
   "Eine Möglichkeit, dies zu umgehen, wäre, die Berechnung in einen Hintergrundprozess oder einen separaten Thread zu verschieben."
   "Sie könnten einen Hintergrund-Worker wie Celery verwenden, um die Optimierungsaufgabe zu verwalten."
 - Wie oft darf gesolvt werden? zb. max 2x pro Woche?

 - Die gerechte Verteilung geht über die max Stunden hinaus wenn zuviele MA benötigt werden und zu wenige Stunden eingegeben wurden?
 -------------------------------

 - Jeder MA muss vor dem Solven eingegeben haben, wann er arbeiten kann. Auch wenn es alles 0 sind.


Prio 2:
 - start_time und end_time zwei und drei noch implementieren
 - der Admin kann auch die Kosten der MA, wenn er will, eintragen. 

"""

class ORAlgorithm:
    def __init__(self, dp: DataProcessing):
        self.current_user_id = dp.current_user_id           # 100     
        self.user_availability = dp.user_availability       # 101
        self.opening_hours = dp.opening_hours               # 102
        self.laden_oeffnet = dp.laden_oeffnet               # 103
        self.laden_schliesst = dp.laden_schliesst           # 104
        self.binary_availability = dp.binary_availability   # 105
        self.company_shifts = dp.company_shifts             # 106
        self.weekly_hours = dp.weekly_hours                 # 107
        self.employment_lvl = dp.employment_lvl             # 108
        self.time_req = dp.time_req                         # 109    
        self.user_employment = dp.user_employment           # 110
        self.solver_requirements = dp.solver_requirements   # 111
        self.week_timeframe = dp.week_timeframe             # 112
        self.hour_devider = dp.hour_devider                 # 113

        # Attribute der Methode "create_variables"
        self.mitarbeiter = None                             # 1
        self.verfügbarkeit = {}                             # 2
        self.kosten = None                                  # 3
        self.max_zeit = None                                # 4
        self.min_zeit = None                                # 5
        self.max_time_week = None                           # 6   
        #self.weekly_hours (* self.hour_devider)            # 7
        self.calc_time = None                               # 8
        self.employment_lvl_exact = []                      # 9
        self.employment = []                                # 10
        self.verteilbare_stunden = None                     # 11
        self.gesamtstunden_verfügbarkeit = []               # 12
        self.min_anwesend = []                              # 13
        self.gerechte_verteilung = []                       # 14
        self.fair_distribution = None                       # 15

        self.desired_max_time_day = None
        self.max_time_day = None
        self.desired_min_time_day = None
        self.min_time_day = None

        # Attribute der Methode "solver_selection"
        self.solver = None               

        # Attribute der Methode "define_penalty_costs"
        self.penalty_cost_nb1 = None
        self.penalty_cost_nb2 = None
        self.penalty_cost_nb3_min = None
        self.penalty_cost_nb4_max = None
        self.penalty_cost_nb5_min = None
        self.penalty_cost_nb6_max = None
        self.penalty_cost_nb7 = None
        self.penalty_cost_nb8 = None

        # Attribute der Methode "decision_variables"
        self.x = None
        self.y = None
        self.a = None
        self.s2 = None
        self.s3 = None
        self.c = None 

        # Attribute der Methode "violation_variables"
        self.nb1_violation = {}
        self.nb2_violation = {}
        self.nb3_min_violation = {}
        self.nb4_max_violation = {}
        self.nb5_min_violation = {}
        self.nb6_max_violation = {}
        self.nb7_violation = {}
        self.nb8_violation = {}

        # Attribute der Methode "objective_function"
        self.objective = None

        # Attribute der Methode "solve_problem"
        self.status = None

        # Attribute der Methode "store_solved_data"
        self.mitarbeiter_arbeitszeiten = {}


    def run(self):
        self.create_variables()
        self.show_variables()
        self.pre_check_programmer()
        # self.pre_check_admin()
        self.solver_selection()
        self.define_penalty_costs()
        self.decision_variables()
        self.violation_variables()
        self.objective_function()
        self.constraints()
        self.solve_problem()
        self.store_solved_data()
        self.output_result_excel()
        self.save_data_in_database()


    def create_variables(self):
        """
        Allgemeine Variabeln
        """
        # -- 1 ------------------------------------------------------------------------------------------------------------
        # user_ids Liste, wird als Key in der Ausgabe verwendet
        self.mitarbeiter = [user_id for user_id in self.binary_availability]

        # -- 2 ------------------------------------------------------------------------------------------------------------
        # Aus dem binary_availability dict. die Verfügbarkeits-Informationen ziehen
        for i, (user_id, availabilities) in enumerate(self.binary_availability.items()):
            self.verfügbarkeit[self.mitarbeiter[i]] = []
            for day_availability in availabilities:
                date, binary_list = day_availability
                self.verfügbarkeit[self.mitarbeiter[i]].append(binary_list)

        # -- 3 --
        # Kosten für jeden MA noch gleich, ebenfalls die max Zeit bei allen gleich
        self.kosten = {ma: 100 for ma in self.mitarbeiter}  # Kosten pro Stunde

        # -- 4 ------------------------------------------------------------------------------------------------------------
        key = "desired_max_time_day"
        if key in self.solver_requirements:
            self.desired_max_time_day = self.solver_requirements[key]
        self.desired_max_time_day = self.desired_max_time_day * self.hour_devider

        key = "max_time_day"
        if key in self.solver_requirements:
            self.max_time_day = self.solver_requirements[key]
        self.max_time_day = self.max_time_day * self.hour_devider             
        
        # Es wird weiterhin ein dict generiert, falls in Zukunft die max_time pro MA verschieden wird
        self.max_zeit = {ma: self.desired_max_time_day for ma in self.mitarbeiter}  # Maximale Arbeitszeit pro Tag

        # -- 5 ------------------------------------------------------------------------------------------------------------
        key = "desired_min_time_day"
        if key in self.solver_requirements:
            self.desired_min_time_day = self.solver_requirements[key]
        self.desired_min_time_day = self.desired_min_time_day * self.hour_devider      

        key = "min_time_day"
        if key in self.solver_requirements:
            self.min_time_day = self.solver_requirements[key]
        self.min_time_day = self.min_time_day * self.hour_devider                  

        # Es wird weiterhin ein dict generiert, falls in Zukunft die min_time pro MA verschieden wird
        self.min_zeit = {ma: self.desired_min_time_day for ma in self.mitarbeiter}  # Minimale Arbeitszeit pro Tag

        # -- 6 ------------------------------------------------------------------------------------------------------------
        # Max. Arbeitszeit pro Woche / Arbeitszeit pro Woche
        key = "max_time_week"
        if key in self.solver_requirements:
            self.max_time_week = self.solver_requirements[key]

        self.max_time_week = self.max_time_week * self.hour_devider

        # -- 7 ------------------------------------------------------------------------------------------------------------
        self.weekly_hours = self.weekly_hours * self.hour_devider                    

        # -- 8 ------------------------------------------------------------------------------------------------------------
        # Berechnung der calc_time (Anzahl Tage an denen die MA eingeteilt werden)
        self.calc_time = 7 * self.week_timeframe

        # -- 9 ------------------------------------------------------------------------------------------------------------
        # Empolyment_level aus dem employment_lvl dict in einer Liste speichern (nur MA die berücksichtigt werden)
        # Iterieren Sie über die Schlüssel in binary_availability
        for user_id in self.binary_availability.keys():
            # Prüfen Sie, ob die user_id in employment_lvl vorhanden ist
            if user_id in self.employment_lvl:
                # Fügen Sie den entsprechenden employment_lvl Wert zur Liste hinzu
                self.employment_lvl_exact.append(self.employment_lvl[user_id])
        # self.employment_lvl_exact = [1, 0.8, 0.8, 0.6, 0.6] # Damit die Liste noch selbst manipuliert werden kann.

        # -- 10 ------------------------------------------------------------------------------------------------------------
        # Iteration of the key within binary_availability
        for user_id in self.binary_availability.keys():
            if user_id in self.user_employment:
                self.employment.append(self.user_employment[user_id])
        # self.employment = ["Perm", "Temp", "Temp", "Temp", "Temp"] # selbst manipuliert

        # -- 11 ------------------------------------------------------------------------------------------------------------
        # verteilbare Stunden (Wieviele Mannstunden benötigt die Firma im definierten Zeitraum)
        self.verteilbare_stunden = 0
        for date in self.time_req:
            for hour in self.time_req[date]:
                self.verteilbare_stunden += self.time_req[date][hour]
                self.verteilbare_stunden = self.verteilbare_stunden

        # -- 12 ------------------------------------------------------------------------------------------------------------
        for key in self.binary_availability:
            gesamt_stunden = sum(sum(day_data[1]) for day_data in self.binary_availability[key])
            self.gesamtstunden_verfügbarkeit.append(gesamt_stunden)

        # -- 13 ------------------------------------------------------------------------------------------------------------
        # Eine Liste mit den min. anwesendheiten der MA wird erstellt
        for _, values in sorted(self.time_req.items()):
            self.min_anwesend.append(list(values.values()))

        # -- 14 ------------------------------------------------------------------------------------------------------------
        # Eine Liste mit den Stunden wie sie gerecht verteilt werden
        list_gesamtstunden = []
        for i in range(len(self.mitarbeiter)):
            if self.gesamtstunden_verfügbarkeit[i] > self.weekly_hours * self.week_timeframe:
                arbeitsstunden_MA = self.employment_lvl_exact[i] * self.weekly_hours * self.week_timeframe
            else:
                arbeitsstunden_MA = self.employment_lvl_exact[i] * self.gesamtstunden_verfügbarkeit[i]
            list_gesamtstunden.append(int(arbeitsstunden_MA))
        print("list_gesamtstunden: ", list_gesamtstunden)

        # Berechnung der Arbeitsstunden für Perm und Temp Mitarbeiter
        total_hours_assigned = 0
        temp_employees = []
        self.gerechte_verteilung = [0 for _ in range(len(self.mitarbeiter))]  # Initialisiere die Liste mit Platzhaltern
        print("1. self.gerechte_verteilung: ", self.gerechte_verteilung)
        for i in range(len(self.mitarbeiter)):
            if self.employment[i] == "Perm":
                allocated_hours = self.employment_lvl_exact[i] * self.weekly_hours * self.week_timeframe
                total_hours_assigned += allocated_hours
                self.gerechte_verteilung[i] = round(allocated_hours + 0.5)
            else:
                temp_employees.append(i)
        print("2. self.gerechte_verteilung: ", self.gerechte_verteilung)

        remaining_hours = self.verteilbare_stunden - total_hours_assigned
        print("remaining_hours: ", remaining_hours)
        for i in temp_employees:
            temp_hours = remaining_hours * (self.employment_lvl_exact[i] / sum(self.employment_lvl_exact[j] for j in temp_employees))
            self.gerechte_verteilung[i] = round(temp_hours + 0.5)
        print("3. self.gerechte_verteilung: ", self.gerechte_verteilung)

        # Wenn die Rundung dazu geführt hat, dass total_hours_assigned die verteilbare_stunden überschreitet, werden die Stunden für Temp-Mitarbeiter angepasst
        total_hours_assigned = sum(self.gerechte_verteilung)
        print("remaining_hours2: ", remaining_hours)
        if total_hours_assigned > self.verteilbare_stunden:
            # Sortieren der Temp-Mitarbeiter nach zugeteilten Stunden in absteigender Reihenfolge
            temp_employees.sort(key=lambda i: self.gerechte_verteilung[i], reverse=True)
            print(temp_employees)
            # Überschüssigen Stunden von den Temp-Mitarbeitern abziehen, beginnend mit demjenigen mit den meisten Stunden
            for i in temp_employees:
                if total_hours_assigned == self.verteilbare_stunden:
                    break
                self.gerechte_verteilung[i] -= 1
                total_hours_assigned -= 1
        print("4. self.gerechte_verteilung: ", self.gerechte_verteilung)       

        # -- 15 ------------------------------------------------------------------------------------------------------------
        # Toleranz der gerechten Verteilung
        key = "fair_distribution"
        if key in self.solver_requirements:
            self.fair_distribution = self.solver_requirements[key]
        self.fair_distribution = self.fair_distribution / 100      # Prozentumrechnung


    def show_variables(self):
        """
        Wenn die Methode aktiviert wird, werden alle Attribute geprintet
        """
        # Attribute aus DataProcessing
        print("100. self.current_user_id: ", self.current_user_id) 
        print("101. self.user_availability: ", self.user_availability) 
        print("102. self.opening_hours: ", self.opening_hours) 
        print("103. self.laden_oeffnet: ", self.laden_oeffnet) 
        print("104. self.laden_schliesst: ", self.laden_schliesst) 
        print("105. self.binary_availability: ", self.binary_availability) 
        print("106. self.company_shifts: ", self.company_shifts) 
        print("107. self.weekly_hours: ", self.weekly_hours)
        print("108. self.employment_lvl: ", self.employment_lvl) 
        print("109. self.time_req: ", self.time_req) 
        print("110. user_employment: ", self.user_employment) 
        print("111. solver_requirements: ", self.solver_requirements)
        print("112. week_timeframe: ", self.week_timeframe)
        print("113. self.hour_devider: ", self.hour_devider)
        print()
        
        print("Attribute der Methode create_variables:")
        # Attribute der Methode "create_variables"
        print("1. self.mitarbeiter: ", self.mitarbeiter)
        print("2. self.verfügbarkeit: ")
        for key, value in self.verfügbarkeit.items():
            print("MA_id: ", key)
            print("Wert: ", value)
        print("3. self.kosten: ", self.kosten)
        print("4. self.max_zeit: ", self.max_zeit)
        print("5. self.min_zeit: ", self.min_zeit)
        print("6. self.max_time_week: ", self.max_time_week)
        print("7. self.weekly_hours: ", self.weekly_hours)
        print("8. self.calc_time: ", self.calc_time)
        print("9. self.empolyment_lvl_exact: ", self.employment_lvl_exact)
        print("10. self.employment: ", self.employment)
        print("11. self.verteilbare_stunden: ", self.verteilbare_stunden)
        print("12. self.gesamtstunden_verfügbarkeit: ", self.gesamtstunden_verfügbarkeit)
        print("13. self.min_anwesend: ", self.min_anwesend)
        print("14. self.gerechte_verteilung: ", self.gerechte_verteilung)
        print("15. self.fair_distribution: ", self.fair_distribution)


    def pre_check_programmer(self):
        """ 
        Vorüberprüfungen für den Programmierer
        """
        # Attribute der Methode "create_variables"
        # -- 1 -- 
        assert isinstance(self.mitarbeiter, list), "self.mitarbeiter sollte eine Liste sein"
        assert all(isinstance(ma, int) for ma in self.mitarbeiter), "Alle Elemente in self.mitarbeiter sollten Ganzzahlen sein"

        # -- 2 -- 
        assert isinstance(self.verfügbarkeit, dict), "self.verfügbarkeit sollte ein Wörterbuch sein"
        assert all(isinstance(val, list) for val in self.verfügbarkeit.values()), "Alle Werte in self.verfügbarkeit sollten Listen sein"
        assert len(self.verfügbarkeit) == len(self.mitarbeiter), "self.verfügbarkeit und self.mitarbeiter sollten die gleiche Länge haben"

        # -- 3 -- 
        assert isinstance(self.kosten, dict), "self.kosten sollte ein Wörterbuch sein"
        assert all(isinstance(kost, (int, float)) for kost in self.kosten.values()), "Alle Werte in self.kosten sollten Ganzzahlen oder Gleitkommazahlen sein"

        # -- 4 -- 
        assert isinstance(self.max_zeit, dict), "self.max_zeit sollte ein Wörterbuch sein"
        assert all(isinstance(zeit, (int, float)) for zeit in self.max_zeit.values()), "Alle Werte in self.max_zeit sollten Ganzzahlen oder Gleitkommazahlen sein"

        # -- 5 -- 
        assert isinstance(self.min_zeit, dict), "self.min_zeit sollte ein Wörterbuch sein"
        assert all(isinstance(zeit, (int, float)) for zeit in self.min_zeit.values()), "Alle Werte in self.min_zeit sollten Ganzzahlen oder Gleitkommazahlen sein"

        # -- 6 -- 
        assert isinstance(self.max_time_week, (int, float)), "self.max_time_week sollte eine Ganzzahl oder eine Gleitkommazahl sein"

        # -- 7 -- 
        assert isinstance(self.calc_time, int), "self.calc_time sollte eine Ganzzahl sein"

        # -- 8 -- 
        assert isinstance(self.employment_lvl_exact, list), "self.employment_lvl_exact sollte eine Liste sein"
        assert all(isinstance(level, (int, float)) for level in self.employment_lvl_exact), "Alle Elemente in self.employment_lvl_exact sollten Ganzzahlen oder Gleitkommazahlen sein"

        # -- 9 -- 
        assert isinstance(self.employment, list), "self.employment sollte eine Liste sein"
        assert all(isinstance(emp, str) for emp in self.employment), "Alle Elemente in self.employment sollten Zeichenketten sein"

        # -- 10 -- 
        assert isinstance(self.verteilbare_stunden, (int, float)), "self.verteilbare_stunden sollte eine Ganzzahl oder eine Gleitkommazahl sein"

        # -- 11 -- 
        assert isinstance(self.gesamtstunden_verfügbarkeit, list), "self.gesamtstunden_verfügbarkeit sollte eine Liste sein"
        assert all(isinstance(stunde, (int, float)) for stunde in self.gesamtstunden_verfügbarkeit), "Alle Elemente in self.gesamtstunden_verfügbarkeit sollten Ganzzahlen oder Gleitkommazahlen sein"

        # -- 12 -- 
        assert isinstance(self.min_anwesend, list), "self.min_anwesend sollte eine Liste sein"
        assert all(isinstance(val, list) for val in self.min_anwesend), "Alle Elemente in self.min_anwesend sollten Listen sein"


    def pre_check_admin(self):
        # Wenn die einzelnen Überprüfungen nicht standhalten, wird ein ValueError ausgelöst und jeweils geprintet, woran das Problem liegt. 
        # Später soll der Admin genau eine solche Meldung angezeigt bekommen.
        
        """
        ---------------------------------------------------------------------------------------------------------------
        1. Überprüfen ob die "Perm" Mitarbeiter mind. self.weekly_hours Stunden einplant haben
        ---------------------------------------------------------------------------------------------------------------
        """
        for i in range(len(self.mitarbeiter)):
            if self.employment[i] == "Perm": 
                sum_availability_perm = 0
                for j in range(self.calc_time):
                    for k in range(len(self.verfügbarkeit[self.mitarbeiter[i]][j])):
                        sum_availability_perm += self.verfügbarkeit[self.mitarbeiter[i]][j][k]
                if sum_availability_perm < self.weekly_hours:
                    raise ValueError(f"Fester Mitarbeiter mit ID {self.mitarbeiter[i]} hat nicht genügend Stunden geplant.")

        """
        ---------------------------------------------------------------------------------------------------------------
        2. Haben die MA mindestens die anzahl Stunden von gerechte_verteilung eingegeben?
        ---------------------------------------------------------------------------------------------------------------
        """
        errors = []

        for index, ma_id in enumerate(self.mitarbeiter):
            if self.employment[index] == 'Temp':
                total_hours_week = sum(sum(self.verfügbarkeit[ma_id][day]) for day in range(self.calc_time))
                if total_hours_week < self.gerechte_verteilung[index]:
                    errors.append(
                        f"Temp-Mitarbeiter {ma_id} hat nur {total_hours_week} Stunden in der Woche eingetragen. "
                        f"Das ist weniger als die erforderliche Gesamtstundenzahl von {self.gerechte_verteilung[index]} Stunden."
                    )
        if errors:
            raise ValueError("Folgende Fehler wurden gefunden:\n" + "\n".join(errors))

        """
        ---------------------------------------------------------------------------------------------------------------
        3. Haben alle MA zusammen genug Stunden eingegeben, um die verteilbaren Stunden zu erreichen?
        ---------------------------------------------------------------------------------------------------------------
        """
        total_hours_available = sum(self.gesamtstunden_verfügbarkeit)
        toleranz = 1 # Wenn man möchte, das die eingegebenen Stunden der MA höher sein müssen als die verteilbaren_stunden

        if total_hours_available < self.verteilbare_stunden * toleranz:
            raise ValueError(f"Die Mitarbeiter haben insgesamt nicht genug Stunden eingegeben, um die verteilbaren Stunden zu erreichen. Benötigte Stunden: {self.verteilbare_stunden}, eingegebene Stunden: {total_hours_available}, Toleranz: {toleranz}")

        """
        ---------------------------------------------------------------------------------------------------------------
        4. Ist zu jeder notwendigen Zeit (self.min_anwesend) die mindestanzahl Mitarbeiter verfügbar?
        ---------------------------------------------------------------------------------------------------------------
        """
        for i in range(len(self.min_anwesend)):  # Für jeden Tag in der Woche
            for j in range(len(self.min_anwesend[i])):  # Für jede Stunde am Tag
                if sum([self.verfügbarkeit[ma][i][j] for ma in self.mitarbeiter]) < self.min_anwesend[i][j]:
                    raise ValueError(f"Es sind nicht genügend Mitarbeiter verfügbar zur notwendigen Zeit (Tag {i+1}, Stunde {j+1}).")

        """
        ---------------------------------------------------------------------------------------------------------------
        5. Können die MA die min. Zeit täglich erreichen? Wenn 0 Stunden eingegeben wurden, läuft es durch!
        ---------------------------------------------------------------------------------------------------------------
        """
        errors = []
        for ma in self.mitarbeiter:
            for day in range(self.calc_time):
                total_hours = sum(self.verfügbarkeit[ma][day])
                if 0 < total_hours < self.min_time_day:
                    errors.append(
                        f"Mitarbeiter {ma} hat am Tag {day+1} nur {total_hours/4} Stunden eingetragen. "
                        f"Das ist weniger als die Mindestarbeitszeit von {self.min_zeit[ma]/4} Stunden."
                    )
        if errors:
            raise ValueError("Folgende Fehler wurden gefunden:\n" + "\n".join(errors))

        """
        ---------------------------------------------------------------------------------------------------------------
        6. Ist die min. Zeit pro Tag so klein, dass die Stunden in der gerechten Verteilung nicht erfüllt werden können?
        ---------------------------------------------------------------------------------------------------------------
        """



        """
        ---------------------------------------------------------------------------------------------------------------
        7. Ist die Toleranz der gerechten Verteilung zu klein gewählt? --> Evtl. die Bedingung weich machen!
        ---------------------------------------------------------------------------------------------------------------
        """


    def solver_selection(self):
        """
        Anwahl des geeigneten Solvers
        # GLOP = Simplex Verfahren
        # CBC =  branch-and-bound- und branch-and-cut-Verfahren
        # SCIP = Framework für die Lösung gemischt-ganzzahliger Programmierungsproblem
        # GLPK = Vielzahl von Algorithmen, einschließlich des Simplex-Verfahrens und des branch-and-bound-Verfahrens
        """
        self.solver = pywraplp.Solver.CreateSolver('SCIP')
        # self.solver.SetNumThreads(4) # Auf mehreren Kernen gleichzeitig arbeiten
        # self.solver.SetTimeLimit(20000)  # Zeitlimit auf 20 Sekunden (in Millisekunden)
        # self.solver.SetSolverSpecificParametersAsString("limits/gap=0.01") # Wenn der gap kleiner 1% ist, bricht der Solver ab


    def define_penalty_costs(self):
        """
        Definiere Strafkosten für weiche Nebenbedingungen
        """
        # Strafkosten für jede NB
        penalty_values = {
            "nb1": {0: 100, 1: 150, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb2": {0: 100, 1: 150, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb3": {0: 100, 1: 150, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb4": {0: 100, 1: 150, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb5": {0: 100, 1: 150, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb6": {0: 100, 1: 150, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb7": {0: 100, 1: 150, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb8": {0: 100, 1: 150, 2: 250, 3: 400 , 4: 600, 5: 10000}
        }

        # Mapping für die entsprechenden Namen der Klassenattribute
        penalty_cost_names = {
            "nb1": "penalty_cost_nb1",
            "nb2": "penalty_cost_nb2",
            "nb3": "penalty_cost_nb3_min",
            "nb4": "penalty_cost_nb4_max",
            "nb5": "penalty_cost_nb5_min",
            "nb6": "penalty_cost_nb6_max",
            "nb7": "penalty_cost_nb7",
            "nb8": "penalty_cost_nb8"
        }

        # Setze die Strafkosten für jede NB basierend auf dem Dictionary
        for key, values in penalty_values.items():
            if key in self.solver_requirements:
                nb_value = self.solver_requirements[key]
                if nb_value in values:
                    setattr(self, penalty_cost_names[key], values[nb_value])
                    print(f"{penalty_cost_names[key].upper()}: {getattr(self, penalty_cost_names[key])}")
                else:
                    print(f"Zahl für {penalty_cost_names[key].upper()} wurde nicht gefunden")
            else:
                print(f"{penalty_cost_names[key].upper()} nicht in solver_requirements gefunden")


    def decision_variables(self):
        """ 
        Entscheidungsvariabeln 
        # solver.NumVar() <-- kontinuierliche Variabeln
        # solver.BoolVar() <-- boolesche Variabeln
        # solver.IntVar() <-- Int Variabeln
        """
        # Arbeitsvariable
        self.x = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):  # Für jeden Tag der calc_time
                for k in range(len(self.verfügbarkeit[i][j])):  # Für jede Stunde des Tages, an dem die Firma geöffnet ist
                    self.x[i, j, k] = self.solver.IntVar(0, 1, f'x[{i}, {j}, {k}]') # Variabeln können nur die Werte 0 oder 1 annehmen

        # Arbeitsblockvariable
        self.y = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):  # Für jeden Tag der calc_time
                for k in range(len(self.verfügbarkeit[i][j])):  # Für jede Stunde des Tages, an dem die Firma geöffnet ist
                    self.y[i, j, k] = self.solver.IntVar(0, 1, f'y[{i}, {j}, {k}]') # Variabeln können nur die Werte 0 oder 1 annehmen

        # Arbeitstagvariable
        self.a = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):  # Für jeden Tag der calc_time
                self.a[i, j] = self.solver.BoolVar(f'a[{i}, {j}]') # Variablen können nur die Werte 0 oder 1 annehmen

        # Schichtvariable Woche (2-Schicht) 
        self.s2 = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):  # Für jeden Tag der calc_time
                self.s2[i, j] = self.solver.IntVar(0, 1, f's2[{i}, {j}]') # Variabeln können nur die Werte 0 oder 1 annehmen

        # Schichtvariable Woche  (3-Schicht) 
        self.s3 = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):  # Für jeden Tag der calc_time
                self.s3[i, j] = self.solver.IntVar(0, 2, f's3[{i}, {j}]') # Variabeln können nur die Werte 0, 1 oder 2 annehmen

        # Gleiche Schichten innerhalb 2 und 4 Wochen  -- IN BEARBEITUNG 21.08.23 --
        self.c = {}
        for i in self.mitarbeiter:
            for j in range(7, self.calc_time):  # Für jeden Tag ab der 2. Woche bis calc_time
                self.c[i, j] = self.solver.IntVar(0, 1, f'c[{i}, {j}]') # Variabeln können nur die Werte 0 oder 1 annehmen


    def violation_variables(self):
        """
        Verletzungsvariabeln

        Definiere Variablen für Nebenbedingungsverletzungen
        self.solver.NumVar(0, self.solver.infinity() <-- Von 0 bis unendlich. für infinity kann man auch eine Zahl einsetzen
        """
        # NB1 violation variable - Mindestanzahl MA zu jeder Stunde an jedem Tag anwesend 
        for j in range(self.calc_time):
            for k in range(len(self.verfügbarkeit[self.mitarbeiter[0]][j])):
                self.nb1_violation[j, k] = self.solver.NumVar(0, self.solver.infinity(), f'nb1_violation[{j}, {k}]')

        # NB2 violation variable - Max. Arbeitszeit pro Woche
        """
        Die diff's beschreiben die Differenz zwischen der gewünschten min und max Arbeitszeiten pro Woche und der maximalen min und max. Arbeitszeiten pro Woche.
        Diese Differenz ist die maximale Anzahl an Verstössen (in den weichen NBs) welche der Solver tätigen darf.
        """
        diff_1 = self.max_time_week - self.weekly_hours
        print("Differenz max Time Week:", diff_1)
        self.nb2_violation = {ma: {} for ma in self.mitarbeiter}
        # Das verschachtelte Dictionary mit Verletzungsvariablen für jede Woche füllen
        for ma in self.mitarbeiter:
            for week in range(1, self.week_timeframe + 1):
                self.nb2_violation[ma][week] = self.solver.NumVar(0, diff_1, f'nb2_violation[{ma}][{week}]')

        # NB3 Mindestarbeitszeit Verletzungsvariable
        diff_2 = self.desired_min_time_day - self.min_time_day    
        print("Differenz min Time day:", diff_2)
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.nb3_min_violation[i, j] = self.solver.NumVar(0, diff_2, 'nb3_min_violation[%i,%i]' % (i, j))

        # NB4 Höchstarbeitszeit Verletzungsvariable
        diff_3 = self.max_time_day - self.desired_max_time_day
        print("Differenz max Time day:", diff_3)
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.nb4_max_violation[i, j] = self.solver.NumVar(0, diff_3, 'nb4_max_violation[%i,%i]' % (i, j))

        # NB5 Mindestarbeitszeit Verletzungsvariable
        for i in self.mitarbeiter:
            self.nb5_min_violation[i] = [self.solver.NumVar(0, self.solver.infinity(), f'nb5_min_violation[{i}][{week}]') for week in range(1, self.week_timeframe + 1)] 

        # NB6 Höchstarbeitszeit Verletzungsvariable
        for i in self.mitarbeiter:
            self.nb6_max_violation[i] = [self.solver.NumVar(0, self.solver.infinity(), f'nb6_max_violation[{i}][{week}]') for week in range(1, self.week_timeframe + 1)]

        # NB7 Innerhalb einer Woche die gleiche Schicht - Verletzungsvariable
        for i in self.mitarbeiter:
            for j in range(7):
                self.nb7_violation[i, j] = self.solver.NumVar(0, self.solver.infinity(), 'nb7_violation[%i,%i]' % (i, j))
        
        # NB8 Innerhalb der zweiten / vierten Woche die gleiche Schicht - Verletzungsvariable
        for i in self.mitarbeiter:
            for j in range(7, self.calc_time):
                self.nb8_violation[i, j] = self.solver.NumVar(0, self.solver.infinity(), 'nb8_violation[%i,%i]' % (i, j))
        

    def objective_function(self):
        """
        Zielfunktion
        """
        self.objective = self.solver.Objective()

        # Kosten MA minimieren
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[i][j])):
                    # Die Kosten werden multipliziert
                    self.objective.SetCoefficient(self.x[i, j, k], self.kosten[i])

        # Kosten Weiche NB1
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                    for k in range(len(self.verfügbarkeit[i][j])):
                        self.objective.SetCoefficient(self.nb1_violation[j, k], self.penalty_cost_nb1)

        # Kosten Weiche NB2
        for i in self.mitarbeiter:
            for week in range(1, self.week_timeframe + 1):
                self.objective.SetCoefficient(self.nb2_violation[i][week], self.penalty_cost_nb2)

        # Kosten für Weiche NB3 Mindestarbeitszeit Verletzung
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.objective.SetCoefficient(self.nb3_min_violation[i, j], self.penalty_cost_nb3_min)

        # Kosten für Weiche NB4 Höchstarbeitszeit Verletzung
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.objective.SetCoefficient(self.nb4_max_violation[i, j], self.penalty_cost_nb4_max)

        # Kosten für Weiche NB5 Mindestarbeitszeit Verletzung
        for i in self.mitarbeiter:
            for week in range(1, self.week_timeframe + 1):
                self.objective.SetCoefficient(self.nb5_min_violation[i][week-1], self.penalty_cost_nb5_min) # Der 0te Wert der Liste, in welchem [week = 1] ist

        # Kosten für Weiche NB6 Höchstarbeitszeit Verletzung
        for i in self.mitarbeiter:
            for week in range(1, self.week_timeframe + 1):
                self.objective.SetCoefficient(self.nb6_max_violation[i][week-1], self.penalty_cost_nb6_max)

        # Kosten für Weiche NB7 "Innerhalb einer Woche immer gleiche Schichten"
        for i in self.mitarbeiter:
            for j in range(7):
                self.objective.SetCoefficient(self.nb7_violation[i, j], self.penalty_cost_nb7)

        # Kosten für Weiche NB8 "Innerhalb der zweiten Woche immer gleiche Schichten"
        for i in self.mitarbeiter:
            for j in range(7, self.calc_time):
                self.objective.SetCoefficient(self.nb8_violation[i, j], self.penalty_cost_nb8)

        # Es wird veruscht, eine Kombination von Werten für die x[i, j, k] zu finden, die die Summe kosten[i]*x[i, j, k] minimiert + weiche NBs            
        self.objective.SetMinimization()


    def constraints(self):
        """
        Beschränkungen / Nebenbedingungen
        # (Die solver.Add() Funktion nimmt eine Bedingung als Argument und fügt sie dem Optimierungsproblem hinzu.)
        """
        # -------------------------------------------------------------------------------------------------------
        # HARTE NB -- NEU 08.08.2023 --
        # NB 0 - Variable a ist 1, wenn der Mitarbeiter an einem Tag arbeitet, sonst 0
        # -------------------------------------------------------------------------------------------------------
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                # sum_x ist die Summe an der ein MA am Tag arbeiten kann
                sum_x = self.solver.Sum(self.x[i, j, k] for k in range(len(self.verfügbarkeit[i][j])))
                self.solver.Add(self.a[i, j] >= sum_x * 1.0 / (len(self.verfügbarkeit[i][j]) + 1))
                self.solver.Add(self.a[i, j] <= sum_x)


        # -------------------------------------------------------------------------------------------------------
        # HARTE NB
        # NB 1 - MA nur einteilen, wenn er verfügbar ist. 
        # -------------------------------------------------------------------------------------------------------
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[i][j])):
                    self.solver.Add(self.x[i, j, k] <= self.verfügbarkeit[i][j][k])


        # -------------------------------------------------------------------------------------------------------
        # HARTE NB
        # NB 2 - Mindestanzahl MA zu jeder Stunde an jedem Tag anwesend
        # -------------------------------------------------------------------------------------------------------
        for j in range(self.calc_time):
            for k in range(len(self.verfügbarkeit[self.mitarbeiter[0]][j])):  # Wir nehmen an, dass alle Mitarbeiter die gleichen Öffnungszeiten haben
                self.solver.Add(self.solver.Sum([self.x[i, j, k] for i in self.mitarbeiter]) >= self.min_anwesend[j][k])
        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB -- NEU 26.07.2023 --
        # NB 2 - Mindestanzahl MA zu jeder Stunde an jedem Tag anwesend 
        # ***** Weiche Nebenbedingung 1 *****
        # -------------------------------------------------------------------------------------------------------
        for j in range(self.calc_time):
            for k in range(len(self.verfügbarkeit[self.mitarbeiter[0]][j])):  # Wir nehmen an, dass alle Mitarbeiter die gleichen Öffnungszeiten haben
                self.solver.Add(self.solver.Sum([self.x[i, j, k] for i in self.mitarbeiter]) - self.min_anwesend[j][k] <= self.nb1_violation[j, k])
                

        """
        # HARTE NB
        # NB 3 - Max. Arbeitszeit pro Woche 
        total_hours = {ma: self.solver.Sum([self.x[ma, j, k] for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[ma][j]))]) for ma in self.mitarbeiter}
        for ma in self.mitarbeiter:
            self.solver.Add(total_hours[ma] <= self.weekly_hours)
        """
        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB -- NEU 28.07.2023 -- --> Muss noch genauer überprüft werden ob es funktioniert!
        # NB 3 - Max. Arbeitszeit pro Woche
        # ***** Weiche Nebenbedingung 2 *****
        # -------------------------------------------------------------------------------------------------------
        if self.week_timeframe == 1:
            total_hours = {ma: self.solver.Sum([self.x[ma, j, k] for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[ma][j]))]) for ma in self.mitarbeiter}
            for ma in self.mitarbeiter:
                self.solver.Add(total_hours[ma] - self.weekly_hours <= self.nb2_violation[ma][1])
                
        elif self.week_timeframe == 2:
            total_hours_week1 = {ma: self.solver.Sum([self.x[ma, j, k] for j in range(self.calc_time // 2) for k in range(len(self.verfügbarkeit[ma][j]))]) for ma in self.mitarbeiter}
            total_hours_week2 = {ma: self.solver.Sum([self.x[ma, j, k] for j in range(self.calc_time // 2, self.calc_time) for k in range(len(self.verfügbarkeit[ma][j]))]) for ma in self.mitarbeiter}
            for ma in self.mitarbeiter:
                self.solver.Add(total_hours_week1[ma] - self.weekly_hours <= self.nb2_violation[ma][1])
                self.solver.Add(total_hours_week2[ma] - self.weekly_hours <= self.nb2_violation[ma][2])
                
        elif self.week_timeframe == 4:
            total_hours_week1 = {ma: self.solver.Sum([self.x[ma, j, k] for j in range(self.calc_time // 4) for k in range(len(self.verfügbarkeit[ma][j]))]) for ma in self.mitarbeiter}
            total_hours_week2 = {ma: self.solver.Sum([self.x[ma, j, k] for j in range(self.calc_time // 4, self.calc_time // 2) for k in range(len(self.verfügbarkeit[ma][j]))]) for ma in self.mitarbeiter}
            total_hours_week3 = {ma: self.solver.Sum([self.x[ma, j, k] for j in range(self.calc_time // 2, 3 * self.calc_time // 4) for k in range(len(self.verfügbarkeit[ma][j]))]) for ma in self.mitarbeiter}
            total_hours_week4 = {ma: self.solver.Sum([self.x[ma, j, k] for j in range(3 * self.calc_time // 4, self.calc_time) for k in range(len(self.verfügbarkeit[ma][j]))]) for ma in self.mitarbeiter}
            
            for ma in self.mitarbeiter:
                self.solver.Add(total_hours_week1[ma] - self.weekly_hours <= self.nb2_violation[ma][1])
                self.solver.Add(total_hours_week2[ma] - self.weekly_hours <= self.nb2_violation[ma][2])
                self.solver.Add(total_hours_week3[ma] - self.weekly_hours <= self.nb2_violation[ma][3])
                self.solver.Add(total_hours_week4[ma] - self.weekly_hours <= self.nb2_violation[ma][4])

     
        """
        # HARTE NB
        # NB 4 - Min. und Max. Arbeitszeit pro Tag
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                # Prüfen, ob der Mitarbeiter an diesem Tag arbeiten kann (z.B. [0, 1, 1] = sum(2))

                # Wenn diese Bedingung nicht erfüllt ist, kann die min und max Zeit verletzt werden! Ändern

                if sum(self.verfügbarkeit[i][j]) >= self.min_zeit[i]:
                    sum_hour = self.solver.Sum(self.x[i, j, k] for k in range(len(self.verfügbarkeit[i][j])))
                    # Es ist nötig, das die min und die max Zeit implementiert ist. 
                    self.solver.Add(sum_hour >= self.min_zeit[i] * self.a[i, j])
                    # NB 4.1 - Die Arbeitszeit eines Mitarbeiters an einem Tag kann nicht mehr als die maximale Arbeitszeit pro Tag betragen
                    self.solver.Add(sum_hour <= self.max_zeit[i] * self.a[i, j])
        """
        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB -- NEU 31.07.2023 --
        # NB 4 - Min. und Max. Arbeitszeit pro Tag
        # ***** Weiche Nebenbedingung 3 und 4 *****
        # -------------------------------------------------------------------------------------------------------
        for i in self.mitarbeiter:
            for j in range(self.calc_time):

                # Wenn die if Bedingung auskommentiert wird, dann wird die min und max Zeit im gleichen Masse verteilt, funktioniert aber noch nicht!

                # if sum(self.verfügbarkeit[i][j]) >= self.min_zeit[i]:
                    sum_hour = self.solver.Sum(self.x[i, j, k] for k in range(len(self.verfügbarkeit[i][j])))

                    
                    # Prüfen, ob die Summe der Arbeitsstunden kleiner als die Mindestarbeitszeit ist
                    self.solver.Add(sum_hour - self.min_zeit[i] * self.a[i, j] >= -self.nb3_min_violation[i, j])
                    self.solver.Add(self.nb3_min_violation[i, j] >= 0)

                    
                    # Prüfen, ob die Summe der Arbeitsstunden größer als die maximale Arbeitszeit ist
                    self.solver.Add(sum_hour - self.max_zeit[i] * self.a[i, j] <= self.nb4_max_violation[i, j])
                    self.solver.Add(self.nb4_max_violation[i, j] >= 0)


        # -------------------------------------------------------------------------------------------------------
        # HARTE NB
        # NB 5 - Anzahl Arbeitsblöcke
        # -------------------------------------------------------------------------------------------------------
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                # Überprüfen, ob der Betrieb an diesem Tag geöffnet ist
                if self.opening_hours[j] > 0:
                    # Für die erste Stunde des Tages
                    self.solver.Add(self.y[i, j, 0] >= self.x[i, j, 0])
                    # Für die restlichen Stunden des Tages
                    for k in range(1, len(self.verfügbarkeit[i][j])):
                        self.solver.Add(self.y[i, j, k] >= self.x[i, j, k] - self.x[i, j, k-1])
                    # Die Summe der y[i, j, k] für einen bestimmten Tag j sollte nicht größer als 1 sein
                    self.solver.Add(self.solver.Sum(self.y[i, j, k] for k in range(len(self.verfügbarkeit[i][j]))) <= 1)
        

        # -------------------------------------------------------------------------------------------------------
        # HARTE NB
        # NB 6 - Verteilungsgrad MA
        # -------------------------------------------------------------------------------------------------------
        verteilungsstunden = {ma: self.solver.Sum([self.x[ma, j, k] for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[ma][j]))]) for ma in self.mitarbeiter}
        for i, ma in enumerate(self.mitarbeiter):
            lower_bound = self.gerechte_verteilung[i] * (1 - self.fair_distribution)
            upper_bound = self.gerechte_verteilung[i] * (1 + self.fair_distribution)
            self.solver.Add(verteilungsstunden[ma] <= upper_bound)
            self.solver.Add(verteilungsstunden[ma] >= lower_bound)


        """
        # HARTE NB
        # NB 7 - Feste Mitarbeiter zu employement_level fest einplanen
        total_hours = {ma: self.solver.Sum([self.x[ma, j, k] for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[ma][j]))]) for ma in self.mitarbeiter}
        for i, ma in enumerate(self.mitarbeiter):
            if self.employment[i] == "Perm": 
                self.solver.Add(total_hours[ma] == self.weekly_hours)
        """
        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB -- NEU 08.08.23 --
        # NB 7 - "Perm" Mitarbeiter zu employement_level fest einplanen
        # ***** Weiche Nebenbedingung 5 und 6 *****
        # -------------------------------------------------------------------------------------------------------
        """
        Alte NB ohne die Wochenrücksicht:
        total_hours = {ma: self.solver.Sum(self.x[ma, j, k] for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[ma][j]))) for ma in self.mitarbeiter}
        for i, ma in enumerate(self.mitarbeiter):
            if self.employment[i] == "Perm": 
                # Prüfen, ob die Gesamtstunden kleiner als die vorgegebenen Arbeitsstunden sind (Unterschreitung)
                self.solver.Add(total_hours[ma] - self.weekly_hours <= self.nb5_min_violation[ma])
                self.solver.Add(self.nb5_min_violation[ma] >= 0)

                # Prüfen, ob die Gesamtstunden größer als die vorgegebenen Arbeitsstunden sind (Überschreitung)
                self.solver.Add(self.weekly_hours - total_hours[ma] <= -self.nb6_max_violation[ma])
                self.solver.Add(self.nb6_max_violation[ma] >= 0)
        """
        for i, ma in enumerate(self.mitarbeiter):
            if self.employment[i] == "Perm":
                for week in range(1, self.week_timeframe + 1):
                    week_start = (week - 1) * (self.calc_time // self.week_timeframe)
                    week_end = week * (self.calc_time // self.week_timeframe)
                    print(f'Week start for employee {ma} in week {week}: {week_start}')
                    print(f'Week end for employee {ma} in week {week}: {week_end}')

                    total_hours_week = self.solver.Sum(self.x[ma, j, k] for j in range(week_start, week_end) for k in range(len(self.verfügbarkeit[ma][j])))
                    print(f'Total hours for employee {ma} in week {week}: {total_hours_week}')

                    # Prüfen, ob die Gesamtstunden kleiner als die vorgegebenen Arbeitsstunden sind (Unterschreitung)
                    self.solver.Add(total_hours_week - self.weekly_hours <= self.nb5_min_violation[ma][week - 1])
                    self.solver.Add(self.nb5_min_violation[ma][week - 1] >= 0)

                    # Prüfen, ob die Gesamtstunden größer als die vorgegebenen Arbeitsstunden sind (Überschreitung)
                    self.solver.Add(self.weekly_hours - total_hours_week <= -self.nb6_max_violation[ma][week - 1])
                    self.solver.Add(self.nb6_max_violation[ma][week - 1] >= 0)


        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB
        # NB 8 - Innerhalb einer Woche immer gleiche Schichten
        # ***** Weiche Nebenbedingung 7 *****
        # -------------------------------------------------------------------------------------------------------
        if self.company_shifts <= 1:
            pass


        elif self.company_shifts == 2:
            for i in self.mitarbeiter:
                for j in range(7):

                    # Hier noch einbauen, das wenn die Stundenanzahl ungerade ist!!

                    first_shift_hours = self.solver.Sum(self.x[i, j, k] for k in range(0, int(len(self.verfügbarkeit[i][j]) / 2))) # Stunden in der ersten Schicht
                    second_shift_hours = self.solver.Sum(self.x[i, j, k] for k in range(int(len(self.verfügbarkeit[i][j]) / 2), len(self.verfügbarkeit[i][j]))) # Stunden in der zweiten Schicht
                    
                    # Kann 0 oder 1 annehmen
                    delta = self.solver.BoolVar("delta")
                    
                    self.solver.Add(first_shift_hours - second_shift_hours - 1000 * delta <= 0)         # 7 - 2 - 1000 * (0,1) <= 0       
                    self.solver.Add(second_shift_hours - first_shift_hours - 1000 * (1 - delta) <= 0)   # 2 - 7 - 1000 * (1 - (0,1)) <= 0    
                    
                    # Hilfsvariable mit s2[i, j] verknüpfen
                    self.solver.Add(self.s2[i, j] == 1 - delta)

            """
            # Harte nb Option zum testen
            for i in self.mitarbeiter:
                for j in range(1, self.calc_time):
                    self.solver.Add(self.s2[i, j] - self.s2[i, j-1] == 0)
            """ 

            # Bedingungen, um sicherzustellen, dass innerhalb einer Woche immer die gleiche Schicht gearbeitet wird
            for i in self.mitarbeiter:
                for j in range(1, 7):
                    diff = self.solver.IntVar(-1, 1, "diff")
                    
                    # Setzen Sie diff gleich der Differenz
                    self.solver.Add(diff == self.s2[i, j] - self.s2[i, j-1])
         
                    # Bedingungen für den "absoluten Wert"
                    self.solver.Add(self.nb7_violation[i, j] >= diff)
                    self.solver.Add(self.nb7_violation[i, j] >= -diff)


        elif self.company_shifts == 3:
            for i in self.mitarbeiter:
                for j in range(7):
                    total_len = len(self.verfügbarkeit[i][j])
                    third_shift_len = total_len // 3
                    first_shift_len = third_shift_len + (total_len % 3) // 2
                    second_shift_len = third_shift_len + (total_len % 3) - (total_len % 3) // 2

                    first_shift_hours = self.solver.Sum(self.x[i, j, k] for k in range(0, first_shift_len))
                    second_shift_hours = self.solver.Sum(self.x[i, j, k] for k in range(first_shift_len, first_shift_len + second_shift_len))
                    third_shift_hours = self.solver.Sum(self.x[i, j, k] for k in range(first_shift_len + second_shift_len, total_len))

                    delta1 = self.solver.BoolVar(f"delta1_{i}_{j}")
                    delta2 = self.solver.BoolVar(f"delta2_{i}_{j}")
                    delta3 = self.solver.BoolVar(f"delta3_{i}_{j}")

                    M = 1000

                    self.solver.Add(first_shift_hours >= second_shift_hours - M * (1 - delta1))
                    self.solver.Add(first_shift_hours >= third_shift_hours - M * (1 - delta1))

                    self.solver.Add(second_shift_hours >= first_shift_hours - M * (1 - delta2))
                    self.solver.Add(second_shift_hours >= third_shift_hours - M * (1 - delta2))

                    self.solver.Add(third_shift_hours >= first_shift_hours - M * (1 - delta3))
                    self.solver.Add(third_shift_hours >= second_shift_hours - M * (1 - delta3))

                    # Sicherstellen, dass nur eine Schicht ausgewählt wird
                    self.solver.Add(delta1 + delta2 + delta3 == 1)

                    # Hilfsvariable mit s3[i, j] verknüpfen
                    self.solver.Add(self.s3[i, j] == 0 * delta1 + 1 * delta2 + 2 * delta3)

            """
            # Harte Bedingung, dass innerhalb einer Woche immer die gleiche Schicht gearbeitet wird
            for i in self.mitarbeiter:
                for j in range(1, self.calc_time):
                    self.solver.Add(self.s3[i, j] - self.s3[i, j-1] == 0)
            """
            
            # Bedingungen, um sicherzustellen, dass innerhalb einer Woche immer die gleiche Schicht gearbeitet wird
            for i in self.mitarbeiter:
                for j in range(1, 7):
                    diff = self.solver.IntVar(-2, 2, "diff") # Unterschied kann -2, -1, 0, 1 oder 2 sein
                    
                    # Setzen Sie diff gleich der Differenz
                    self.solver.Add(diff == self.s3[i, j] - self.s3[i, j-1])

                    # Bedingungen für den "absoluten Wert"
                    self.solver.Add(self.nb7_violation[i, j] >= diff)
                    self.solver.Add(self.nb7_violation[i, j] >= -diff)
                    self.solver.Add(self.nb7_violation[i, j] <= 1)  # Die Verletzung sollte maximal 1 betragen
            

        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB
        # NB 9 - Wechselnde Schichten innerhalb von 2 und 4 Wochen
        # ***** Weiche Nebenbedingung 8 *****
        # -------------------------------------------------------------------------------------------------------
        
        # 2 Wochen + 2-Schicht # --------------------------------------------------------------------------------
        if self.week_timeframe == 2:
            if self.company_shifts == 2:
                for i in self.mitarbeiter:
                    # Anzahl der Tage in der ersten Woche, an denen in der ersten bzw. zweiten Schicht gearbeitet wurde
                    first_week_first_shift_days = self.solver.Sum(self.s2[i, j] for j in range(7))
                    first_week_second_shift_days = 7 - first_week_first_shift_days

                    # Hilfsvariable, um die Schicht der ersten Woche festzulegen
                    first_week_shift = self.solver.BoolVar("first_week_shift")

                    # Wenn die Anzahl der Tage in der ersten Schicht größer ist, setzen Sie first_week_shift auf 1
                    self.solver.Add(first_week_first_shift_days - first_week_second_shift_days - 1000 * first_week_shift <= 0)
                    self.solver.Add(first_week_second_shift_days - first_week_first_shift_days - 1000 * (1 - first_week_shift) <= 0)

                    for j in range(7, 14):
                        self.solver.Add(self.c[i, j] == 1 - first_week_shift)


                    # In der zweiten Woche muss der Mitarbeiter in der entgegengesetzten Schicht arbeiten
                    for j in range(7, 14):
                        # Summe der Stunden in der ersten Schicht in der zweiten Woche
                        first_shift_hours_second_week = self.solver.Sum(self.x[i, j, k] for k in range(0, int(len(self.verfügbarkeit[i][j]) / 2)))
                        second_shift_hours_second_week = self.solver.Sum(self.x[i, j, k] for k in range(int(len(self.verfügbarkeit[i][j]) / 2), len(self.verfügbarkeit[i][j])))

                        # Kann 0 oder 1 annehmen
                        delta_2 = self.solver.BoolVar("delta_2")
                        
                        self.solver.Add(first_shift_hours_second_week - second_shift_hours_second_week - 1000 * delta_2 <= 0)      
                        self.solver.Add(second_shift_hours_second_week - first_shift_hours_second_week - 1000 * (1 - delta_2) <= 0)  
                        
                        # Hilfsvariable mit s2[i, j] verknüpfen
                        self.solver.Add(self.s2[i, j] == 1 - delta_2)
                        
                        # Harte Nebenbedingung
                        # self.solver.Add(self.s2[i, j] == self.c[i, j])

                # Verletzungsvariable erhöhen, wenn die Schicht in der zweiten Woche nicht der entgegengesetzten Schicht entspricht
                for i in self.mitarbeiter:
                    for j in range(7, 14):
                        diff = self.solver.IntVar(-1, 1, f"diff_{i}_{j}")
                        self.solver.Add(diff == self.s2[i, j] - self.c[i, j])

                        self.solver.Add(self.nb8_violation[i, j] >= diff)
                        self.solver.Add(self.nb8_violation[i, j] >= -diff)

        
        # 2 Wochen + 3-Schicht # --------------------------------------------------------------------------------
        """
        if self.week_timeframe == 2 and self.company_shifts == 3:
            for i in self.mitarbeiter:
                first_week_first_shift_days_list = []
                first_week_second_shift_days_list = []
                
                for j in range(7):
                    # Hilfsvariablen für die Schicht an Tag j
                    is_first_shift = self.solver.BoolVar(f"is_first_shift_{i}_{j}")
                    is_second_shift = self.solver.BoolVar(f"is_second_shift_{i}_{j}")
                    
                    # Verknüpfen der Hilfsvariablen mit s3
                    self.solver.Add(is_first_shift - self.s3[i, j] == 0)
                    self.solver.Add(is_second_shift - (self.s3[i, j] - 1) == 0)

                    # Zu Listen hinzufügen, um später zu summieren
                    first_week_first_shift_days_list.append(is_first_shift)
                    first_week_second_shift_days_list.append(is_second_shift)

                # Summe der Tage in der jeweiligen Schicht berechnen
                first_week_first_shift_days = self.solver.Sum(first_week_first_shift_days_list)
                first_week_second_shift_days = self.solver.Sum(first_week_second_shift_days_list)
                first_week_third_shift_days = 7 - first_week_first_shift_days - first_week_second_shift_days
                

                first_week_shift_0 = self.solver.BoolVar("first_week_shift_0")
                first_week_shift_1 = self.solver.BoolVar("first_week_shift_1")
                first_week_shift_2 = self.solver.BoolVar("first_week_shift_2")

                self.solver.Add(first_week_first_shift_days >= first_week_second_shift_days + first_week_third_shift_days - 1000 * first_week_shift_0)
                self.solver.Add(first_week_second_shift_days >= first_week_first_shift_days + first_week_third_shift_days - 1000 * first_week_shift_1)
                self.solver.Add(first_week_third_shift_days >= first_week_first_shift_days + first_week_second_shift_days - 1000 * first_week_shift_2)

                self.solver.Add(first_week_shift_0 + first_week_shift_1 + first_week_shift_2 == 1)

                for j in range(7, 14):  # zweite Woche
                    self.solver.Add(self.c[i, j] == 1 * first_week_shift_0 + 2 * first_week_shift_1 + 0 * first_week_shift_2)
                
                for j in range(7, 14):  # zweite Woche
                    first_shift_hours_second_week = self.solver.Sum(self.x[i, j, k] for k in range(0, total_len // 3))
                    second_shift_hours_second_week = self.solver.Sum(self.x[i, j, k] for k in range(total_len // 3, 2 * total_len // 3))
                    third_shift_hours_second_week = self.solver.Sum(self.x[i, j, k] for k in range(2 * total_len // 3, total_len))

                    delta2_1 = self.solver.BoolVar(f"delta2_1_{i}_{j}")
                    delta2_2 = self.solver.BoolVar(f"delta2_2_{i}_{j}")
                    delta2_3 = self.solver.BoolVar(f"delta2_3_{i}_{j}")

                    M = 1000

                    self.solver.Add(first_shift_hours_second_week >= second_shift_hours_second_week + third_shift_hours_second_week - M * (1 - delta2_1))
                    self.solver.Add(second_shift_hours_second_week >= first_shift_hours_second_week + third_shift_hours_second_week - M * (1 - delta2_2))
                    self.solver.Add(third_shift_hours_second_week >= first_shift_hours_second_week + second_shift_hours_second_week - M * (1 - delta2_3))

                    self.solver.Add(delta2_1 + delta2_2 + delta2_3 == 1)

                    self.solver.Add(self.s3[i, j] == 0 * delta2_1 + 1 * delta2_2 + 2 * delta2_3)

                    # Harte Nebenbedingung
                    self.solver.Add(self.s3[i, j] == self.c[i, j])

            
            # Bedingungen, um sicherzustellen, dass in der zweiten Woche die Schicht um eins verschoben wird
            for i in self.mitarbeiter:
                for j in range(7, 14):  # Zweite Woche
                    diff = self.solver.IntVar(-2, 2, f"diff_{i}_{j}")  # Der Unterschied kann -2, -1, 0, 1 oder 2 sein
                    
                    # Setzen Sie diff gleich der Differenz
                    self.solver.Add(diff == self.s3[i, j] - self.c[i, j])
                    
                    # Bedingungen für den zulässigen Wert
                    shift_up_condition = self.solver.BoolVar(f"shift_up_condition_{i}_{j}")
                    shift_down_condition = self.solver.BoolVar(f"shift_down_condition_{i}_{j}")

                    # Setzt shift_up_condition auf 1, wenn diff == 1, sonst auf 0
                    self.solver.Add(diff <= 1 + 2*(1 - shift_up_condition))
                    self.solver.Add(diff >= 1 - 2*(1 - shift_up_condition))

                    # Setzt shift_down_condition auf 1, wenn diff == -2, sonst auf 0
                    self.solver.Add(diff <= -2 + 2*(1 - shift_down_condition))
                    self.solver.Add(diff >= -2 - 2*(1 - shift_down_condition))

                    # Es sollte entweder eine Aufwärtsverschiebung oder eine Abwärtsverschiebung geben
                    self.solver.Add(shift_up_condition + shift_down_condition == 1)

                    # Bedingungen für den "absoluten Wert"
                    self.solver.Add(self.nb8_violation[i, j] >= diff)
                    self.solver.Add(self.nb8_violation[i, j] >= -diff)
                    self.solver.Add(self.nb8_violation[i, j] <= 1)  # Die Verletzung sollte maximal 1 betragen
            """

        








        # 4 Wochen + 2-Schicht ----------------------------------------------------------------------------------
        if self.week_timeframe == 4:
            if self.company_shifts == 2:
                for i in self.mitarbeiter:
                    # Anzahl der Tage in der ersten Woche, an denen in der ersten bzw. zweiten Schicht gearbeitet wurde
                    first_week_first_shift_days = self.solver.Sum(self.s2[i, j] for j in range(7))
                    first_week_second_shift_days = 7 - first_week_first_shift_days

                    # Hilfsvariable, um die Schicht der ersten Woche festzulegen
                    first_week_shift = self.solver.BoolVar("first_week_shift")

                    # Wenn die Anzahl der Tage in der ersten Schicht größer ist, setzen Sie first_week_shift auf 1
                    self.solver.Add(first_week_first_shift_days - first_week_second_shift_days - 1000 * first_week_shift <= 0)
                    self.solver.Add(first_week_second_shift_days - first_week_first_shift_days - 1000 * (1 - first_week_shift) <= 0)

                    for week in range(2, 5): # Beginne bei Woche 2
                        for j in range((week-1)*7, week*7): # Beginne bei Tag 0 der aktuellen Woche und gehe bis zum letzten Tag der Woche
                            if week % 2 == 0:
                                self.solver.Add(self.c[i, j] == 1 - first_week_shift)
                            else:
                                self.solver.Add(self.c[i, j] == first_week_shift)


                    for week in range(2, 5): # Wochen 2 bis 4
                        for j in range((week-1)*7, week*7):
                            # Summe der Stunden in der ersten Schicht in der aktuellen Woche
                            first_shift_hours_current_week = self.solver.Sum(self.x[i, j, k] for k in range(0, int(len(self.verfügbarkeit[i][j]) / 2)))
                            second_shift_hours_current_week = self.solver.Sum(self.x[i, j, k] for k in range(int(len(self.verfügbarkeit[i][j]) / 2), len(self.verfügbarkeit[i][j])))

                            # Kann 0 oder 1 annehmen
                            delta_2 = self.solver.BoolVar("delta_2")

                            self.solver.Add(first_shift_hours_current_week - second_shift_hours_current_week - 1000 * delta_2 <= 0)
                            self.solver.Add(second_shift_hours_current_week - first_shift_hours_current_week - 1000 * (1 - delta_2) <= 0)

                            # Hilfsvariable mit s2[i, j] verknüpfen
                            self.solver.Add(self.s2[i, j] == 1 - delta_2)

                            # Harte Nebenbedingung
                            # self.solver.Add(self.s2[i, j] == self.c[i, j])
                
                # Verletzungsvariable erhöhen, wenn die Schicht in den Wochen 2, 3 und 4 nicht der entgegengesetzten Schicht entspricht
                for i in self.mitarbeiter:
                    for j in range(7, 28):
                        diff = self.solver.IntVar(-1, 1, f"diff_{i}_{j}")
                        self.solver.Add(diff == self.s2[i, j] - self.c[i, j])

                        self.solver.Add(self.nb8_violation[i, j] >= diff)
                        self.solver.Add(self.nb8_violation[i, j] >= -diff)
                
                
        # 4 Wochen + 3-Schicht ----------------------------------------------------------------------------------
        if self.week_timeframe == 4:
            if self.company_shifts == 3:
                for i in self.mitarbeiter:
                    pass



       

    def solve_problem(self):
        """
        Problem lösen und Kosten ausgeben
        """
        self.solver.EnableOutput(False)
        self.status = self.solver.Solve()


        # --------------------------------------------------------------------------------------
        # Die Werte von s3 printen
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                # Drucken Sie den Wert von s3[i, j]
                print(f"s3[{i}][{j}] =", self.s3[i, j].solution_value())

        # Die Werte von c printen
        for i in self.mitarbeiter:
            for j in range(7, self.calc_time):
                # Drucken Sie den Wert von c[i, j]
                print(f"c[{i}][{j}] =", self.c[i, j].solution_value())
        # --------------------------------------------------------------------------------------


        # Kosten für die Einstellung von Mitarbeitern
        hiring_costs = sum(self.kosten[i] * self.x[i, j, k].solution_value() for i in self.mitarbeiter for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[i][j])))

        # Strafen für die Verletzung der weichen Nebenbedingungen
        nb1_penalty_costs = sum(self.penalty_cost_nb1 * self.nb1_violation[j, k].solution_value() for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[self.mitarbeiter[0]][j])))
        nb2_penalty_costs = sum(self.penalty_cost_nb2 * self.nb2_violation[i][week].solution_value() for i in self.mitarbeiter for week in range(1, self.week_timeframe + 1))
        nb3_min_penalty_costs = sum(self.penalty_cost_nb3_min * self.nb3_min_violation[i, j].solution_value() for i in self.mitarbeiter for j in range(self.calc_time))
        nb4_max_penalty_costs = sum(self.penalty_cost_nb4_max * self.nb4_max_violation[i, j].solution_value() for i in self.mitarbeiter for j in range(self.calc_time))
        nb5_min_penalty_costs = sum(self.penalty_cost_nb5_min * self.nb5_min_violation[i][week - 1].solution_value() for i in self.mitarbeiter for week in range(1, self.week_timeframe + 1))
        nb6_max_penalty_costs = sum(self.penalty_cost_nb6_max * self.nb6_max_violation[i][week - 1].solution_value() for i in self.mitarbeiter for week in range(1, self.week_timeframe + 1))
        nb7_penalty_costs = sum(self.penalty_cost_nb7 * self.nb7_violation[i, j].solution_value() for i in self.mitarbeiter for j in range(7))
        nb8_penalty_costs = sum(self.penalty_cost_nb8 * self.nb8_violation[i, j].solution_value() for i in self.mitarbeiter for j in range(7, self.calc_time))

        # Kosten der einzelnen NBs ausgeben
        print('Kosten Einstellung von Mitarbeitern:', hiring_costs)
        print('Kosten Weiche NB1 (Mindestanzahl MA zu jeder Stunde an jedem Tag anwesend):', nb1_penalty_costs)
        print('Kosten Weiche NB2 (Max. Arbeitszeit pro Woche "Temp" MA):', nb2_penalty_costs)
        print('Kosten Weiche NB3 (Min. Arbeitszeit pro Tag):', nb3_min_penalty_costs)
        print('Kosten Weiche NB4 (Max. Arbeitszeit pro Tag):', nb4_max_penalty_costs)
        print('Kosten Weiche NB5 (Unterschreitung der festen Mitarbeiter zu employment_level):', nb5_min_penalty_costs)
        print('Kosten Weiche NB6 (Überschreitung der festen Mitarbeiter zu employment_level):', nb6_max_penalty_costs)
        print('Kosten Weiche NB7 (Immer die gleiche Schicht in einer Woche):', nb7_penalty_costs)
        print('Kosten Weiche NB8 (Immer die gleiche Schicht zweite Woche):', nb8_penalty_costs)
        print('Gesamtkosten:', self.objective.Value())


    def store_solved_data(self):
        """
        mitarbeiter_arbeitszeiten Attribut befüllen
        """
        if self.status == pywraplp.Solver.OPTIMAL or self.status == pywraplp.Solver.FEASIBLE:
            for i in self.mitarbeiter:
                self.mitarbeiter_arbeitszeiten[i] = []
                for j in range(self.calc_time):
                    arbeitszeit_pro_tag = []
                    for k in range(len(self.verfügbarkeit[i][j])):
                        arbeitszeit_pro_tag.append(int(self.x[i, j, k].solution_value()))
                    self.mitarbeiter_arbeitszeiten[i].append(arbeitszeit_pro_tag)
            print(self.mitarbeiter_arbeitszeiten)

        if self.status == pywraplp.Solver.OPTIMAL:
            print("Optimale Lösung gefunden.")
        elif self.status == pywraplp.Solver.FEASIBLE:
            print("Mögliche Lösung gefunden.")
        elif self.status == pywraplp.Solver.INFEASIBLE:
            print("Problem ist unlösbar.")
        elif self.status == pywraplp.Solver.UNBOUNDED:
            print("Problem ist unbeschränkt.")
        elif self.status == pywraplp.Solver.NOT_SOLVED:
            print("Solver hat das Problem nicht gelöst.")
        else:
            print("Unbekannter Status.")


    def output_result_excel(self):
        """
        Excel ausgabe
        """
        data = self.mitarbeiter_arbeitszeiten

        # Legen Sie die Füllungen fest
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Festlegen der Schriftgröße
        font = Font(size=10)
        header_font = Font(size=5)  # Schriftgröße für die Spaltentitel

        # Erstellen Sie ein Workbook
        wb = Workbook()
        ws = wb.active

        # Schreiben Sie die Überschriften
        headers = ["user_id"]
        for i in range(1, len(data[list(data.keys())[0]]) + 1):
            headers.extend(["T{}, {}:{}".format(i, j+8, k*15) for j in range(10) for k in range(4)])
            headers.append(' ')
        headers.append("Total Hours") 
        ws.append(headers)

        # Ändern der Schriftgröße der Spaltentitel
        for cell in ws[1]:
            cell.font = header_font

        # Schreiben Sie die Daten
        for idx, (ma, days) in enumerate(data.items(), start=2):
            row = [ma]
            for day in days:
                quarter_hours = [h for h in day]
                row.extend(quarter_hours)
                row.append(' ')  # Fügt eine leere Spalte nach jedem Tag hinzu
            ws.append(row)
            total_hours_col = len(headers)  # Wir verwenden die Anzahl der Überschriften, um die richtige Spalte zu erhalten
            ws.cell(row=idx, column=total_hours_col, value=f"=SUM(B{idx}:{get_column_letter(total_hours_col-1)}{idx})")


        # Farben auf Basis der Zellenwerte festlegen und Schriftgröße für den Rest des Dokuments
        for row in ws.iter_rows(min_row=2, values_only=False):
            for cell in row[1:]:
                if cell.value == 1:
                    cell.fill = green_fill
                elif cell.value == 0:
                    cell.fill = red_fill
                cell.font = font

        # Ändern der Spaltenbreite
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 3:
                adjusted_width = 3
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Speichern Sie das Workbook
        wb.save("Einsatzplan.xlsx")


    def save_data_in_database(self):
        """ 
        Diese Methode speichert die berechneten Arbeitszeiten in der Datenbank 
        """
        with app.app_context():
            for user_id, days in self.mitarbeiter_arbeitszeiten.items(): # Durch mitarbeiter_arbeitszeiten durchitterieren
                print(f"Verarbeite Benutzer-ID: {user_id}")

                # Benutzer aus der Datenbank abrufen
                user = User.query.get(user_id)
                print(user)
                if not user:
                    print(f"Kein Benutzer gefunden mit ID: {user_id}")
                    continue

                for day_index, day in enumerate(days):
                    # Wir gehen davon aus, dass der erste Tag im 'self.user_availability' das Startdatum ist
                    date = self.user_availability[user_id][0][0] + datetime.timedelta(days=day_index)
                    print("DATE: ", date)

                    # Löschen der jeweiligen Tage
                    Timetable.query.filter_by(email=user.email, date=date).delete()
                    db.session.commit()
                    

                    # Hier unterteilen wir den Tag in Schichten, basierend auf den Zeiten, zu denen der Mitarbeiter arbeitet
                    shifts = []
                    start_time_index = None
                    for time_index in range(len(day)):
                        if day[time_index] == 1 and start_time_index is None:
                            start_time_index = time_index
                        elif day[time_index] == 0 and start_time_index is not None:
                            shifts.append((start_time_index, time_index))
                            start_time_index = None
                    
                    if start_time_index is not None:
                        shifts.append((start_time_index, len(day)))

                    print(f"Berechnete Schichten für Benutzer-ID {user_id}, Tag-Index {day_index}: {shifts}")

                    # Divisor bestimmen
                    divisor = 3600 / self.hour_devider

                    for shift_index, (start_time, end_time) in enumerate(shifts):
                        # Ladenöffnungszeit am aktuellen Tag hinzufügen
                        opening_time_in_units = int(self.laden_oeffnet[day_index].total_seconds() * self.hour_devider / 3600)
                        start_time += opening_time_in_units
                        end_time += opening_time_in_units

                        # Neues Timetable-Objekt
                        new_entry = Timetable(
                            id=None,  # ID wird automatisch generiert
                            email=user.email,
                            first_name=user.first_name,
                            last_name=user.last_name,
                            company_name=user.company_name,
                            date=date,
                            start_time=datetime.datetime.combine(date, datetime.time(hour=int(start_time // self.hour_devider), minute=int((start_time % self.hour_devider) * 60 / self.hour_devider))),
                            end_time=datetime.datetime.combine(date, datetime.time(hour=int(end_time // self.hour_devider), minute=int((end_time % self.hour_devider) * 60 / self.hour_devider))),
                            start_time2=None,
                            end_time2=None,
                            start_time3=None,
                            end_time3=None,
                            created_by=self.current_user_id,
                            changed_by=self.current_user_id,
                            creation_timestamp=datetime.datetime.now()
                        )

                        # new_entry der Datenbank hinzufügen
                        db.session.add(new_entry)

            # Änderungen in der Datenbank speichern
            db.session.commit()