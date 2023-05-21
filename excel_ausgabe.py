import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Daten
data = {'MA1': [[0, 0, 0, 0, 0, 0, 0, 0, 0, 1], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]], 'MA2': [[0, 0, 0, 1, 0, 1, 1, 0, 0, 0], [0, 0, 0, 1, 1, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 1, 0, 0, 0, 0, 0, 0]], 'MA3': [[0, 0, 0, 0, 1, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 1, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 1, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]], 'MA4': [[1, 1, 1, 0, 0, 0, 0, 1, 1, 0], [0, 0, 0, 0, 0, 0, 0, 0, 1, 0], [0, 0, 0, 1, 1, 0, 1, 0, 1, 0], [0, 0, 0, 1, 0, 0, 0, 1, 0, 0], [0, 0, 0, 0, 1, 0, 0, 1, 0, 0]], 'MA5': [[0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [1, 1, 1, 0, 0, 1, 1, 0, 0, 1], [1, 1, 1, 0, 0, 1, 0, 1, 0, 1], [1, 1, 1, 0, 1, 0, 1, 0, 1, 1], [1, 1, 1, 0, 0, 1, 1, 0, 1, 1]]}


# Legen Sie die Füllungen fest
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Festlegen der Schriftgröße
font = Font(size=10)
header_font = Font(size=5)  # Schriftgröße für die Spaltentitel

# Erstellen Sie ein Workbook
wb = Workbook()
ws = wb.active

# Schreiben Sie die Überschriften
headers = ["MA"]
for i in range(1, len(max(data.values(), key=len)) + 7):
    headers.extend(["T{},h{}".format(i, j+8) for j in range(10)])
ws.append(headers)

# Ändern der Schriftgröße der Spaltentitel
for cell in ws[1]:
    cell.font = header_font

# Schreiben Sie die Daten
for ma, days in data.items():
    row = [ma]
    for day in days:
        if day:
            row.extend(day)
        else:
            row.extend([None]*9)  # Für Tage ohne Stunden
    ws.append(row)

# Farben auf Basis der Zellenwerte festlegen und Schriftgröße für den Rest des Dokuments
for row in ws.iter_rows(min_row=2, values_only=False):
    for cell in row[1:]:
        if cell.value == 1:
            cell.fill = green_fill
        elif cell.value == 0:
            cell.fill = red_fill
        cell.font = font

# Ändern der Spaltenbreite
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    if adjusted_width > 3:
        adjusted_width = 3
    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Speichern Sie das Workbook
wb.save("Einsatzplan.xlsx")
