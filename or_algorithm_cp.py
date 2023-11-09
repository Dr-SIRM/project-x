import pandas as pd
import time
import datetime
import math
import re
import threading
import os

from app import db, app, timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

from data_processing import DataProcessing
from sqlalchemy import text
from models import Timetable, User, SolverAnalysis
from collections import defaultdict

from ortools.sat.python import cp_model
"""
Der CP-SAT Solver arbeitet mit Boolescher Zufriedenheit (SAT, Satisfiability) 
und Constraint Programming (CP) Techniken, um Diskretheit, Kombinatorik und 
Bedingungen in den Problemen zu behandeln. Er ist speziell für Probleme entwickelt worden, 
bei denen komplexe und kombinatorische Beschränkungen berücksichtigt werden müssen.
"""


"""
To-Do Liste:

 - (erl.) NB6: Max. einen Arbeitsblock pro Tag
 - (erl.) die calc_time soll automatisch errechnet werden
 - (erl.) Als Key oder i für MA soll nicht mehr MA1, MA2 usw stehen, sondern die user_id (zB. 1002)
 - (erl.) Shifts/Employment_level aus der Datenbank ziehen
 - (erl.) auf Viertelstunden wechseln
 - (erl.) Gerechte Verteilung Anpassen, das Stunden von "Perm" Mitarbeiter abgezogen werden
 - (erl.) Weiche NB4 implementieren, hat noch nicht wunschgemäss geklappt
 - (erl.) Den Übergang auf harte und weiche NBs machen? 
 - (erl.) Die gesolvten Daten in der Datenbank speichern
 - (erl.) Daten für Solven in die Datenbank einpflegen (max. Zeit, min. Zeit, Solvingzeitraum, Toleranz für die Stundenverteilung, ...)
 - (erl.) Eine if Anweseiung, wenn der Betrieb an einem Tag geschlossen hat. Dann soll an diesem Tag nicht gesolvet werden
 - (erl.) Stunden Teiler für 1/4, 1/2 und 1h einbauen
 - (erl.) NB10 die weiche fertigbauen
 - (erl.) Opening Hour 2 einbauen
 - (erl.) Shifts aus Solver_req Datenbank rauslöschen und Daten nicht mehr ziehen
 - (erl.) Während des Solvings Daten ziehen --> Fragen gestellt
 - (erl.) Die erste Vorüberprüfung funktioniert noch nicht, da auch ein Perm MA der 60% angestellt ist weekly_hours eingteilt werden muss.
 - (erl) pre_check_admin aus data_processing in or_algorithm einpflegen
 - (erl) gerechte Verteilung angepasst
 - (erl) Wenn min Zeit grösser als gewünschte, dann Fehler -> beheben!
 - (erl) Überprüfen, ob eigegebene werte bei Verletzungsvariabeln wirklich korrekt sind
 - (erl) NB11 und 12 implementieren
 - (erl) Bei den max. Arbeitstagen eine obere schranke einbauen (harte NB 10)
 - (erl) excel_output Daten anderst ziehen
 - (erl) Wenn Montag ist, wird für diese Woche gesolvt statt für nächste Woche
 - (erl) Vorüberprüfungen sauber beschreiben damit es vernünftig angezeigt wird.
 - (erl) get_availability und binaere_liste anpassen 
 - (erl) start_time und end_time zwei und drei noch implementieren
 - (erl) TimeReq in data_processing anpassen, soabld Phu Planung fertiggestellt hat.
 - (erl) MA mit verschiedenen Profilen - Department (Koch, Service, ..)
 - (erl) Alle NB's mit Schichten umbauen.
 

 To-Do's 
 -------------------------------
 - Solvingzeit mit der Zeit automaitsch abbrechen, Datenbank mit Variabeln befüllen und Daten ziehen
 - (*)  Wenn man gar keine time_req eingegeben hat, hällt dann Vorüberprüfung 1 stand?

 - (*) self.subsequent_workingdays_max, self.skills_per_day(1 oder 0) in die Datenbank einpflegen und ziehen
 - self.max_time_week darf niemals grösser als self.weekly hours gewählt werden!
 - self.daily_deployment soll nur 1 oder 2 werden können


 
 --- PRIO 2 ---
 -------------------------------
 - Die gerechte Verteilung geht über die max Stunden hinaus wenn zuviele MA benötigt werden und zu wenige Stunden eingegeben wurden??
 - Solvingzeitraum selbst anwählen können
 - (*) NB9 mit 3 Schichten fertigbauen
 -------------------------------

 
"""

class ORAlgorithm_cp:
    def __init__(self, dp: DataProcessing):
        self.start_date = dp.start_date
        self.end_date = dp.end_date

        self.current_user_id = dp.current_user_id           # 100     
        self.user_availability = dp.user_availability       # 101
        self.opening_hours = dp.opening_hours               # 102
        self.laden_oeffnet = dp.laden_oeffnet               # 103
        self.laden_schliesst = dp.laden_schliesst           # 104
        self.binary_availability = dp.binary_availability   # 105
        self.company_shifts = dp.company_shifts             # 106
        self.weekly_hours = dp.weekly_hours                 # 107
        self.employment_lvl = dp.employment_lvl             # 108
        self.time_req = dp.time_req                         # 109    
        self.user_employment = dp.user_employment           # 110
        self.solver_requirements = dp.solver_requirements   # 111
        self.week_timeframe = dp.week_timeframe             # 112
        self.hour_devider = dp.hour_devider                 # 113
        self.user_names = dp.user_names                     # 114
        self.user_skills = dp.user_skills                   # 115
        self.skills = dp.skills                             # 116
        self.user_holidays = dp.user_holidays               # 117

        # Attribute der Methode "create_variables"
        self.mitarbeiter = None                             # 1
        self.mitarbeiter_s = {}                             # 1.2
        self.verfügbarkeit = {}                             # 2
        self.kosten = None                                  # 3
        self.max_zeit = None                                # 4
        self.min_zeit = None                                # 5
        self.max_time_week = None                           # 6   
        #self.weekly_hours (* self.hour_devider)            # 7
        self.calc_time = None                               # 8
        self.employment_lvl_exact = []                      # 9
        self.employment = []                                # 10
        self.verteilbare_stunden = None                     # 11
        self.gesamtstunden_verfügbarkeit = []               # 12
        self.min_anwesend = {}                              # 13
        self.gerechte_verteilung = []                       # 14
        self.fair_distribution = None                       # 15
        self.subsequent_workingdays = None                  # 16
        self.daily_deployment = None                        # 17
        self.time_per_deployment = None                     # 18
        self.benoetigte_skills = []                         # 19

        self.desired_max_time_day = None
        self.max_time_day = None
        self.desired_min_time_day = None
        self.min_time_day = None


        # Attribute der Methode "solver_selection"
        self.model = None
        self.solver = None
        self.gap_to_stop = None              

        # Attribute der Methode "define_penalty_costs"
        self.penalty_cost_nb1 = None
        self.penalty_cost_nb2 = None
        self.penalty_cost_nb3_min = None
        self.penalty_cost_nb4_max = None
        self.penalty_cost_nb5_min = None
        self.penalty_cost_nb6_max = None
        self.penalty_cost_nb7 = None
        self.penalty_cost_nb8 = None
        self.penalty_cost_nb9 = None
        self.penalty_cost_nb10 = None
        self.penalty_cost_nb11_min = None
        self.penalty_cost_nb12_max = None

        # Attribute der Methode "decision_variables"
        self.x = None
        self.y = None
        self.a = None
        self.a_sum = None
        self.s2 = None
        self.s3 = None
        self.c = None 

        # Attribute der Methode "violation_variables"
        self.nb1_violation = {}
        self.nb2_violation = {}
        self.nb3_min_violation = {}
        self.nb4_max_violation = {}
        self.nb5_min_violation = {}
        self.nb6_max_violation = {}
        self.nb7_violation = {}
        self.nb8_violation = {}
        self.nb9_violation = {}
        self.nb10_violation = {}
        self.nb11_min_violation = {}
        self.nb12_max_violation = {}

        # Attribute der Methode "objective_function"
        self.cost_expressions = []

        # Attribute der Methode "solve_problem"
        self.best_values = []
        self.status = None
        self.solving_time_seconds = None

        # Attribute der Methode "calculate_costs"
        self.violation_nb1 = None
        self.violation_nb2 = None
        self.violation_nb3 = None
        self.violation_nb4 = None
        self.violation_nb5 = None
        self.violation_nb6 = None
        self.violation_nb7 = None
        self.violation_nb8 = None
        self.violation_nb9 = None
        self.violation_nb10 = None
        self.violation_nb11 = None
        self.violation_nb12 = None

        self.hiring_costs = None
        self.nb1_penalty_costs = None
        self.nb2_penalty_costs = None
        self.nb3_min_penalty_costs = None
        self.nb4_max_penalty_costs = None
        self.nb5_min_penalty_costs = None
        self.nb6_max_penalty_costs = None
        self.nb7_penalty_costs = None
        self.nb8_penalty_costs = None
        self.nb9_penalty_costs = None
        self.nb10_penalty_costs = None
        self.nb11_min_penalty_costs = None
        self.nb12_max_penalty_costs = None

        # Attribute der Methode "store_solved_data"
        self.mitarbeiter_arbeitszeiten = {}


    def run(self):
        self.create_variables()
        self.show_variables()
        # self.pre_check_programmer()

    # Diese Methode kann später gelöscht werden, da auf die einzelnen pre_checks zugegriffen wird.
    def pre_check(self):
        results = {
            'checks': []
        }

        results['checks'].append(self.pre_check_1())
        results['checks'].append(self.pre_check_2())
        results['checks'].append(self.pre_check_3())
        results['checks'].append(self.pre_check_4())
        results['checks'].append(self.pre_check_5())
        results['checks'].append(self.pre_check_6())
        # results['checks'].append(self.pre_check_7())
        # results['checks'].append(self.pre_check_8())

        print("Test Check:", results)

        return results


    def run_2(self):
        self.solver_selection()
        self.define_penalty_costs()
        self.decision_variables()
        self.violation_variables()
        self.objective_function()
        self.constraints()
        self.solve_problem()
        self.calculate_costs()
        self.store_solved_data()
        self.output_result_excel()
        self.plot_costs_excel()
        self.save_data_in_database()
        self.save_data_in_database_testing()


    def create_variables(self):
        """
        Allgemeine Variabeln
        """
        # -- 1 ------------------------------------------------------------------------------------------------------------
        # user_ids Liste, wird als Key in der Ausgabe verwendet
        self.mitarbeiter = [user_id for user_id in self.binary_availability]


        # Erstellen Sie ein neues Wörterbuch basierend auf den user_ids in self.binary_availability
        for user_id in self.binary_availability:
            # Überprüfen Sie, ob die user_id in self.user_skills vorhanden ist
            if user_id in self.user_skills:
                # Füge die user_id und die zugehörigen Skills dem neuen Wörterbuch hinzu
                self.mitarbeiter_s[user_id] = self.user_skills[user_id]

        # -- 2 ------------------------------------------------------------------------------------------------------------
        # Aus dem binary_availability dict. die Verfügbarkeits-Informationen ziehen
        for i, (user_id, availabilities) in enumerate(self.binary_availability.items()):
            self.verfügbarkeit[self.mitarbeiter[i]] = []
            for day_availability in availabilities:
                date, binary_list = day_availability
                self.verfügbarkeit[self.mitarbeiter[i]].append(binary_list)

        # -- 3 ------------------------------------------------------------------------------------------------------------
        # Kosten für jeden MA noch gleich, ebenfalls die max Zeit bei allen gleich
        self.kosten = {ma: 100 for ma in self.mitarbeiter}  # Kosten pro Stunde

        # -- 4 ------------------------------------------------------------------------------------------------------------
        key = "desired_max_time_day"
        if key in self.solver_requirements:
            self.desired_max_time_day = self.solver_requirements[key]
        self.desired_max_time_day = self.desired_max_time_day * self.hour_devider

        key = "max_time_day"
        if key in self.solver_requirements:
            self.max_time_day = self.solver_requirements[key]
        self.max_time_day = self.max_time_day * self.hour_devider             
        
        # Es wird weiterhin ein dict generiert, falls in Zukunft die max_time pro MA verschieden wird
        self.max_zeit = {ma: self.desired_max_time_day for ma in self.mitarbeiter}  # Maximale Arbeitszeit pro Tag

        # -- 5 ------------------------------------------------------------------------------------------------------------
        key = "desired_min_time_day"
        if key in self.solver_requirements:
            self.desired_min_time_day = self.solver_requirements[key]
        self.desired_min_time_day = self.desired_min_time_day * self.hour_devider

        key = "min_time_day"
        if key in self.solver_requirements:
            self.min_time_day = self.solver_requirements[key]
        self.min_time_day = self.min_time_day * self.hour_devider                 

        # Es wird weiterhin ein dict generiert, falls in Zukunft die min_time pro MA verschieden wird
        self.min_zeit = {ma: self.desired_min_time_day for ma in self.mitarbeiter}  # Minimale Arbeitszeit pro Tag

        # -- 6 ------------------------------------------------------------------------------------------------------------
        # Max. Arbeitszeit pro Woche / Arbeitszeit pro Woche
        key = "max_time_week"
        if key in self.solver_requirements:
            self.max_time_week = self.solver_requirements[key]

        self.max_time_week = self.max_time_week * self.hour_devider

        # -- 7 ------------------------------------------------------------------------------------------------------------
        self.weekly_hours = self.weekly_hours * self.hour_devider                    

        # -- 8 ------------------------------------------------------------------------------------------------------------
        # Berechnung der calc_time (Anzahl Tage an denen die MA eingeteilt werden)
        self.calc_time = 7 * self.week_timeframe

        # -- 9 ------------------------------------------------------------------------------------------------------------
        # Empolyment_level aus dem employment_lvl dict in einer Liste speichern (nur MA die berücksichtigt werden)
        # Iterieren Sie über die Schlüssel in binary_availability
        for user_id in self.binary_availability.keys():
            # Prüfen Sie, ob die user_id in employment_lvl vorhanden ist
            if user_id in self.employment_lvl:
                # Fügen Sie den entsprechenden employment_lvl Wert zur Liste hinzu
                self.employment_lvl_exact.append(self.employment_lvl[user_id])
        # self.employment_lvl_exact = [1, 0.8, 0.8, 0.6, 0.6] # Damit die Liste noch selbst manipuliert werden kann.

        # -- 10 ------------------------------------------------------------------------------------------------------------
        # Iteration of the key within binary_availability
        for user_id in self.binary_availability.keys():
            if user_id in self.user_employment:
                self.employment.append(self.user_employment[user_id])

        # -- 11 ------------------------------------------------------------------------------------------------------------
        # verteilbare Stunden (Wieviele Mannstunden benötigt die Firma im definierten Zeitraum)

        self.verteilbare_stunden = 0
        # Durchlaufen OrderedDict
        for date, roles in self.time_req.items():
            if isinstance(roles, defaultdict):
                # Durchlaufen jeder Rolle und deren Stundenanforderungen
                for role, hours in roles.items():
                    if isinstance(hours, defaultdict):
                        # Addieren der Anforderungen für jede Stunde
                        for hour, count in hours.items():
                            self.verteilbare_stunden += count

        # -- 12 ------------------------------------------------------------------------------------------------------------
        for key in self.binary_availability:
            gesamt_stunden = sum(sum(day_data[1]) for day_data in self.binary_availability[key])
            self.gesamtstunden_verfügbarkeit.append(gesamt_stunden)

        # -- 13 ------------------------------------------------------------------------------------------------------------
        # Eine Dict mit den min. anwesendheiten der MA wird erstellt

        # Durch alle Elemente in self.time_req iterieren
        for _, day_skills in self.time_req.items():
            for skill, hours in day_skills.items():
                # Überprüfen, ob der Skill bereits im Dictionary vorhanden ist
                if skill not in self.min_anwesend:
                    self.min_anwesend[skill] = []
                
                # Die Stunden für den aktuellen Tag zum Skill hinzufügen
                self.min_anwesend[skill].append(hours)


        # -- 13.5 ------------------------------------------------------------------------------------------------------------
        # Ein dict erstellen mit den Anzahl Urlaubstagen pro User

        # Maximale Anzahl der Ferientage basierend auf den Öffnungstagen berechnen
        max_holidays = sum(1 for hours in self.opening_hours if hours > 0)

        # Leeres Dictionary für die Gesamtanzahl der Urlaubstage pro Benutzer
        self.total_holidays_per_user = {}
        for user_id, dates in self.user_holidays.items():
            # Summe der Urlaubstage für den aktuellen Benutzer initialisieren
            total_holidays = 0
            
            # Alle Urlaubstage für den aktuellen Benutzer addieren
            for _, holidays in dates:
                total_holidays += holidays
            
            # Urlaubstage auf die maximale Anzahl der Öffnungstage limitieren
            total_holidays = min(total_holidays, max_holidays)
            
            # Summe der Urlaubstage dem Benutzer im Dictionary zuweisen
            self.total_holidays_per_user[user_id] = total_holidays


        # -- 14 ------------------------------------------------------------------------------------------------------------
        # gerechte_verteilung -> eine Liste mit den Stunden wie sie gerecht verteilt werden

        # Berechnung der Arbeitsstunden für Perm und Temp Mitarbeiter
        total_hours_assigned = 0
        temp_employees = []
        self.gerechte_verteilung = [0 for _ in range(len(self.mitarbeiter))]  # Initialisiere die Liste mit Platzhaltern
        opening_days = sum(1 for item in self.opening_hours if item > 0) # Anzahl Tage an denen der Betrieb geöffnet ist im Solvingzeitraum
        
        print("1. self.gerechte_verteilung: ", self.gerechte_verteilung)

        # Stunden den Vollzeit Mitarbeiter zuweisen (Wochenarbeitsstunde * Anstellungsgrad * week_timeframe * Ferienzeit Faktor (1 - (holidays_user / opening_days))
        for i, employee in enumerate(self.mitarbeiter):
            if self.employment[i] == "Perm":
                holidays_user = self.total_holidays_per_user[employee] # Anzahl Ferientage jedes "Perm" users
                allocated_hours = (self.employment_lvl_exact[i] * self.weekly_hours * self.week_timeframe) * (1 - (holidays_user / opening_days))
                print(f"Verhältniss von MA {employee} das * gerechnet wird: ", (1 - (holidays_user / opening_days)))
                total_hours_assigned += allocated_hours
                self.gerechte_verteilung[i] = round(allocated_hours + 0.5)
            else:
                temp_employees.append(i)
        print("2. self.gerechte_verteilung: ", self.gerechte_verteilung)

        # Die übrigen Stunden unter den den Teilzeit Mitarbeitern aufteilen
        remaining_hours = self.verteilbare_stunden - total_hours_assigned
        print("remaining_hours: ", remaining_hours)
        for i in temp_employees:
            temp_hours = remaining_hours * (self.employment_lvl_exact[i] / sum(self.employment_lvl_exact[j] for j in temp_employees))
            self.gerechte_verteilung[i] = round(temp_hours + 0.5)
        print("3. self.gerechte_verteilung: ", self.gerechte_verteilung)

        # Wenn die Rundung dazu geführt hat, dass total_hours_assigned die verteilbare_stunden überschreitet, werden die Stunden für Temp-Mitarbeiter angepasst
        total_hours_assigned = sum(self.gerechte_verteilung)
        if total_hours_assigned > self.verteilbare_stunden:
            # Sortieren der Temp-Mitarbeiter nach zugeteilten Stunden in absteigender Reihenfolge
            temp_employees.sort(key=lambda i: self.gerechte_verteilung[i], reverse=True)
            print("Temporäre Mitarbeiter: ", temp_employees)
            # Überschüssigen Stunden von den Temp-Mitarbeitern abziehen, beginnend mit demjenigen mit den meisten Stunden
            for i in temp_employees:
                if total_hours_assigned == self.verteilbare_stunden:
                    break
                self.gerechte_verteilung[i] -= 1
                total_hours_assigned -= 1
        print("4. self.gerechte_verteilung: ", self.gerechte_verteilung)   

        # Neue Prüfung und Anpassung auf Basis von self.gesamtstunden_verfügbarkeit
        # Wenn gerechte Verteilung grösser als Gesamtstundenverfügbarkeit ist, dann den Wert überschreiben
        for i in range(len(self.gerechte_verteilung)):
            if self.gerechte_verteilung[i] > self.gesamtstunden_verfügbarkeit[i]:
                self.gerechte_verteilung[i] = self.gesamtstunden_verfügbarkeit[i]
        print("5. self.gerechte_verteilung: ", self.gerechte_verteilung)  

        # Erneute Überprüfung der gesamten zugewiesenen Stunden
        total_hours_assigned = sum(self.gerechte_verteilung)
        if total_hours_assigned < self.verteilbare_stunden:
            # Erneute Verteilung der verbleibenden Stunden auf die Mitarbeiter
            additional_hours = self.verteilbare_stunden - total_hours_assigned
            print("additional_hours", additional_hours)

            # Verteilung der zusätzlichen Stunden auf die Temp-Mitarbeiter
            while additional_hours > 0 and any(self.gerechte_verteilung[i] < self.gesamtstunden_verfügbarkeit[i] for i in temp_employees):
                for i in temp_employees:
                    # Wenn das Hinzufügen einer Stunde nicht die maximale Verfügbarkeit überschreitet und es noch verbleibende Stunden gibt
                    if self.gerechte_verteilung[i] < self.gesamtstunden_verfügbarkeit[i] and additional_hours > 0:
                        self.gerechte_verteilung[i] += 1
                        additional_hours -= 1
                        total_hours_assigned += 1

        # Abschließende Überprüfung und Ausgabe
        print("6. self.gerechte_verteilung: ", self.gerechte_verteilung)  
        print("Final total_hours_assigned: ", total_hours_assigned)
        # self.gerechte_verteilung = [40, 40]


        # -- 15 ------------------------------------------------------------------------------------------------------------
        # Toleranz der gerechten Verteilung
        key = "fair_distribution"
        if key in self.solver_requirements:
            self.fair_distribution = self.solver_requirements[key]
        self.fair_distribution = self.fair_distribution / 100      # Prozentumrechnung

        # -- 16 ------------------------------------------------------------------------------------------------------------
        # Maximale Arbeitstage in Folge
        key = "subsequent_workingdays"
        if key in self.solver_requirements:
            self.subsequent_workingdays = self.solver_requirements[key]

        # -- 17 ------------------------------------------------------------------------------------------------------------
        # Anzahl Arbeitseinsätze pro Tag
        key = "daily_deployment"
        if key in self.solver_requirements:
            self.daily_deployment = self.solver_requirements[key]

        # -- 18 ------------------------------------------------------------------------------------------------------------
        # Mindeststunden pro Arbeitsblock
        key = "time_per_deployment"
        if key in self.solver_requirements:
            self.time_per_deployment = self.solver_requirements[key]

        # -- 19 ------------------------------------------------------------------------------------------------------------
        # Alle Skills pro Woche die im Zeitraum benötigt werden in einer Liste zusammenfassen. 

        # Initialisieren von self.benoetigte_skills als Liste von leeren Listen für jede Woche
        self.benoetigte_skills = [[] for _ in range(self.week_timeframe)]

        # Durchlaufen aller Skills und deren Verfügbarkeiten
        for skill, verfuegbarkeiten in self.min_anwesend.items():
            # Für jede "Woche" im gegebenen Zeitrahmen
            for i in range(self.week_timeframe):
                # Start- und Endindex für den 7-Tage-Zeitraum festlegen
                start_index = i * 7
                end_index = start_index + 7

                # Durchlaufen der Verfügbarkeiten innerhalb des aktuellen 7-Tage-Zeitraums
                for verfuegbarkeit in verfuegbarkeiten[start_index:end_index]:
                    # Wenn zu irgendeinem Zeitpunkt in dieser Woche eine Verfügbarkeit besteht, Skill hinzufügen
                    if verfuegbarkeit and skill not in self.benoetigte_skills[i]:
                        self.benoetigte_skills[i].append(skill)
                        break  # Innere Schleife beenden und mit dem nächsten Intervall fortfahren, da der Skill bereits hinzugefügt wurde




    def show_variables(self):
        """
        Wenn die Methode aktiviert wird, werden alle Attribute geprintet
        """
        # Attribute aus DataProcessing
        print("100. self.current_user_id: ", self.current_user_id) 
        print("101. self.user_availability: ", self.user_availability) 
        print("102. self.opening_hours: ", self.opening_hours) 
        print("103. self.laden_oeffnet: ", self.laden_oeffnet) 
        print("104. self.laden_schliesst: ", self.laden_schliesst) 
        print("105. self.binary_availability: ", self.binary_availability) 
        print("106. self.company_shifts: ", self.company_shifts) 
        print("107. self.weekly_hours: ", self.weekly_hours)
        print("108. self.employment_lvl: ", self.employment_lvl) 
        print("109. self.time_req: ", self.time_req) 
        print("110. user_employment: ", self.user_employment) 
        print("111. solver_requirements: ", self.solver_requirements)
        print("112. week_timeframe: ", self.week_timeframe)
        print("113. self.hour_devider: ", self.hour_devider)
        print("114. self.user_names: ", self.user_names)
        print("115. self.user_skills: ", self.user_skills)
        print("116. self.skills: ", self.skills)
        print("117. self.user_holidays: ", self.user_holidays)
        print()
        
        print("Attribute der Methode create_variables:")
        # Attribute der Methode "create_variables"
        print("1. self.mitarbeiter: ", self.mitarbeiter)
        print("1.2 self.mitarbeiter_s: ", self.mitarbeiter_s)
        print("2. self.verfügbarkeit: ")
        for key, value in self.verfügbarkeit.items():
            print("MA_id: ", key)
            print("Wert: ", value)
        print("3. self.kosten: ", self.kosten)
        print("4. self.max_zeit: ", self.max_zeit)
        print("5. self.min_zeit: ", self.min_zeit)
        print("6. self.max_time_week: ", self.max_time_week)
        print("7. self.weekly_hours: ", self.weekly_hours)
        print("8. self.calc_time: ", self.calc_time)
        print("9. self.empolyment_lvl_exact: ", self.employment_lvl_exact)
        print("10. self.employment: ", self.employment)
        print("11. self.verteilbare_stunden: ", self.verteilbare_stunden)
        print("12. self.gesamtstunden_verfügbarkeit: ", self.gesamtstunden_verfügbarkeit)
        print("13. self.min_anwesend: ", self.min_anwesend)
        print("13.5 self.total_holidays_per_user: ", self.total_holidays_per_user)
        print("14. self.gerechte_verteilung: ", self.gerechte_verteilung)
        print("15. self.fair_distribution: ", self.fair_distribution)
        print("16. self.subsequent_workingdays: ", self.subsequent_workingdays)
        print("17. self.daily_deployment: ", self.daily_deployment)
        print("18. self.time_per_deployment: ", self.time_per_deployment)
        print("19. self.benoetigte_skills: ", self.benoetigte_skills)


    def pre_check_programmer(self):
        """ 
        Vorüberprüfungen für den Programmierer
        """
        # Attribute der Methode "create_variables"
        # -- 1 -- 
        assert isinstance(self.mitarbeiter, list), "self.mitarbeiter sollte eine Liste sein"
        assert all(isinstance(ma, int) for ma in self.mitarbeiter), "Alle Elemente in self.mitarbeiter sollten Ganzzahlen sein"

        # -- 2 -- 
        assert isinstance(self.verfügbarkeit, dict), "self.verfügbarkeit sollte ein Wörterbuch sein"
        assert all(isinstance(val, list) for val in self.verfügbarkeit.values()), "Alle Werte in self.verfügbarkeit sollten Listen sein"
        assert len(self.verfügbarkeit) == len(self.mitarbeiter), "self.verfügbarkeit und self.mitarbeiter sollten die gleiche Länge haben"

        # -- 3 -- 
        assert isinstance(self.kosten, dict), "self.kosten sollte ein Wörterbuch sein"
        assert all(isinstance(kost, (int, float)) for kost in self.kosten.values()), "Alle Werte in self.kosten sollten Ganzzahlen oder Gleitkommazahlen sein"

        # -- 4 -- 
        assert isinstance(self.max_zeit, dict), "self.max_zeit sollte ein Wörterbuch sein"
        assert all(isinstance(zeit, (int, float)) for zeit in self.max_zeit.values()), "Alle Werte in self.max_zeit sollten Ganzzahlen oder Gleitkommazahlen sein"

        # -- 5 -- 
        assert isinstance(self.min_zeit, dict), "self.min_zeit sollte ein Wörterbuch sein"
        assert all(isinstance(zeit, (int, float)) for zeit in self.min_zeit.values()), "Alle Werte in self.min_zeit sollten Ganzzahlen oder Gleitkommazahlen sein"

        # -- 6 -- 
        assert isinstance(self.max_time_week, (int, float)), "self.max_time_week sollte eine Ganzzahl oder eine Gleitkommazahl sein"

        # -- 7 -- 
        assert isinstance(self.calc_time, int), "self.calc_time sollte eine Ganzzahl sein"

        # -- 8 -- 
        assert isinstance(self.employment_lvl_exact, list), "self.employment_lvl_exact sollte eine Liste sein"
        assert all(isinstance(level, (int, float)) for level in self.employment_lvl_exact), "Alle Elemente in self.employment_lvl_exact sollten Ganzzahlen oder Gleitkommazahlen sein"

        # -- 9 -- 
        assert isinstance(self.employment, list), "self.employment sollte eine Liste sein"
        assert all(isinstance(emp, str) for emp in self.employment), "Alle Elemente in self.employment sollten Zeichenketten sein"

        # -- 10 -- 
        assert isinstance(self.verteilbare_stunden, (int, float)), "self.verteilbare_stunden sollte eine Ganzzahl oder eine Gleitkommazahl sein"

        # -- 11 -- 
        assert isinstance(self.gesamtstunden_verfügbarkeit, list), "self.gesamtstunden_verfügbarkeit sollte eine Liste sein"
        assert all(isinstance(stunde, (int, float)) for stunde in self.gesamtstunden_verfügbarkeit), "Alle Elemente in self.gesamtstunden_verfügbarkeit sollten Ganzzahlen oder Gleitkommazahlen sein"

        # -- 12 -- 
        assert isinstance(self.min_anwesend, list), "self.min_anwesend sollte eine Liste sein"
        assert all(isinstance(val, list) for val in self.min_anwesend), "Alle Elemente in self.min_anwesend sollten Listen sein"


    def pre_check_1(self):
        """
        ---------------------------------------------------------------------------------------------------------------
        1. Vorüberprüfung: Haben Sie in jeder Stunde eingegeben, wieviele Mitarbeiter benötigt werden?
        Überprüfen, ob der Admin zu jeder Öffnungszeitstunde time_req eingegeben hat
        ---------------------------------------------------------------------------------------------------------------
        """
        # Funktion um Zeitwerte umzurechnen ---------------------------------------------------------------------------
        def time_to_int(t):
            # Divisor basierend auf self.hour_devider erzeugen
            divisor = 3600 / self.hour_devider

            # Typ "timedelta"
            if isinstance(t, timedelta):
                total_seconds = t.total_seconds()  # oder t.seconds, je nachdem, welches Verhalten Sie bevorzugen
            # Typ "time"
            elif isinstance(t, time):
                total_seconds = t.hour * 3600 + t.minute * 60 + t.second
            # Typ "int" oder "float" (z.B. zur direkten Eingabe von Sekunden)
            elif isinstance(t, (int, float)):
                total_seconds = t
            else:
                raise ValueError("Invalid input type, must be datetime.timedelta, datetime.time, int or float")
            
            return int(total_seconds / divisor)
        # -------------------------------------------------------------------------------------------------------------

        try:
            fehlende_stunden = []

            # Erzeugt alle möglichen Daten innerhalb des Bereichs
            start_date = datetime.datetime.strptime(self.start_date, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(self.end_date, "%Y-%m-%d").date()
            date_range = [start_date + timedelta(days=x) for x in range((end_date-start_date).days + 1)]

            for current_date in date_range:
                # Wochentag als Index (0 = Montag, 1 = Dienstag, usw.) erhalten
                weekday_index = current_date.weekday()

                # Prüft, ob der Tag ein Arbeitstag ist (Basierend auf den Öffnungszeiten)
                if self.laden_oeffnet[weekday_index] is not None and self.laden_schliesst[weekday_index] is not None:
                    time_req_dict = self.time_req.get(current_date, {})

                    for hour in range(time_to_int(self.laden_schliesst[weekday_index]) - time_to_int(self.laden_oeffnet[weekday_index])):
                        if hour not in time_req_dict:
                            fehlende_stunden.append((current_date, hour))
            if fehlende_stunden:
                error_message_lines = ["Für folgende Zeitfenster haben Sie noch keine Mitarbeiter eingeplant:"]
                for date, hour in fehlende_stunden:
                    error_message_lines.append(f"Datum: {date}, Stunde: {hour}")
                raise ValueError("\n".join(error_message_lines))
                        
            return {"success": True, "name": "Pre-check_1", "message": "All checks are successful!"}
        except ValueError as e:
            return {"success": False, "name": "Pre-check_1", "message": str(e)}


    def pre_check_2(self):
        """
        ---------------------------------------------------------------------------------------------------------------
        2. Vorüberprüfung: Stehen die Vollzeit Mitarbeiter mindestens die Wochenarbeitsstunden zur Verfügung?
        Überprüfen ob die "Perm" Mitarbeiter mind. self.weekly_hours Stunden einplant haben
        ---------------------------------------------------------------------------------------------------------------
        """
        try:
            for i in range(len(self.mitarbeiter)):
                if self.employment[i] == "Perm": 
                    sum_availability_perm = 0
                    for j in range(self.calc_time):
                        for k in range(len(self.verfügbarkeit[self.mitarbeiter[i]][j])):
                            sum_availability_perm += self.verfügbarkeit[self.mitarbeiter[i]][j][k]

                    if sum_availability_perm < self.weekly_hours * self.employment_lvl_exact[i]:
                        raise ValueError(f"Fester Mitarbeiter mit ID {self.mitarbeiter[i]} hat nicht genügend Stunden geplant.")
                    
            return {"success": True, "name": "Pre-check_2", "message": "All checks are successful!"}
        except ValueError as e:
            return {"success": False, "name": "Pre-check_2", "message": str(e)}


    def pre_check_3(self):
        """
        ---------------------------------------------------------------------------------------------------------------
        3. Vorüberprüfung: Haben die Mitarbeiter mindestens die anzahl Stunden welche sie eingeteilt werden eingegeben?
        Haben die MA mindestens die anzahl Stunden von gerechte_verteilung eingegeben?
        ---------------------------------------------------------------------------------------------------------------
        """
        try: 
            errors = []

            for index, ma_id in enumerate(self.mitarbeiter):
                if self.employment[index] == 'Temp':
                    total_hours_week = sum(sum(self.verfügbarkeit[ma_id][day]) for day in range(self.calc_time))
                    if total_hours_week < self.gerechte_verteilung[index]:
                        errors.append(
                            f"Temp-Mitarbeiter {ma_id} hat nur {total_hours_week} Stunden in der Woche eingetragen. "
                            f"Das ist weniger als die erforderliche Gesamtstundenzahl von {self.gerechte_verteilung[index]} Stunden."
                        )
            if errors:
                raise ValueError("Folgende Fehler wurden gefunden:\n" + "\n".join(errors))

            return {"success": True, "name": "Pre-check_3", "message": "All checks are successful!"}
        except ValueError as e:
            return {"success": False, "name": "Pre-check_3", "message": str(e)}

    def pre_check_4(self):
        """
        ---------------------------------------------------------------------------------------------------------------
        4. Vorüberprüfung: Haben alle Mitarbeiter zusammen genug Stunden eingeplant, um ihre Planung zu erfüllen?
        Haben alle MA zusammen genug Stunden eingegeben, um die verteilbaren Stunden zu erreichen?
        ---------------------------------------------------------------------------------------------------------------
        """
        try: 
            total_hours_available = sum(self.gesamtstunden_verfügbarkeit)
            toleranz = 1 # Wenn man möchte, das die eingegebenen Stunden der MA höher sein müssen als die verteilbaren_stunden

            if total_hours_available < self.verteilbare_stunden * toleranz:
                pass
                raise ValueError(f"Die Mitarbeiter haben insgesamt nicht genug Stunden eingegeben, um die verteilbaren Stunden zu erreichen. Benötigte Stunden: {self.verteilbare_stunden}, eingegebene Stunden: {total_hours_available}, Toleranz: {toleranz}")
            
            return {"success": True, "name": "Pre-check_4", "message": "All checks are successful!"}
        except ValueError as e:
            return {"success": False, "name": "Pre-check_4", "message": str(e)}


    def pre_check_5(self):
        """
        ---------------------------------------------------------------------------------------------------------------
        5. Vorüberprüfung: Stehen zu jeder Zeit mindestens die Anzahl Mitarbeiter zur Verfügung, die Sie benötigten?
        Ist zu jeder notwendigen Zeit (self.min_anwesend) die mindestanzahl Mitarbeiter verfügbar?
        ---------------------------------------------------------------------------------------------------------------
        """
        try: 
            for i in range(len(self.min_anwesend)):  # Für jeden Tag in der Woche
                for j in range(len(self.min_anwesend[i])):  # Für jede Stunde am Tag
                    if sum([self.verfügbarkeit[ma][i][j] for ma in self.mitarbeiter]) < self.min_anwesend[i][j]:
                        #raise ValueError(f"Es sind nicht genügend Mitarbeiter verfügbar zur notwendigen Zeit (Tag {i+1}, Stunde {j+1}).")
                        pass
            
            return {"success": True, "name": "Pre-check_5", "message": "All checks are successful!"}
        except ValueError as e:
            return {"success": False, "name": "Pre-check_5", "message": str(e)}
        
        
    def pre_check_6(self):
        """
        ---------------------------------------------------------------------------------------------------------------
        6. Vorüberprüfung: Haben die Mitarbeiter pro Tag mindestens die mindest Areitszeit pro Tag eingegeben? (bei 0 Stunden wird es ignoriert)
        Können die MA die min. Zeit täglich erreichen? Wenn 0 Stunden eingegeben wurden, läuft es durch!
        ---------------------------------------------------------------------------------------------------------------
        """
        try:
            errors = []
            for i, ma in enumerate(self.mitarbeiter):
                for day in range(self.calc_time):
                    total_hours = sum(self.verfügbarkeit[ma][day])
                    if 0 < total_hours < self.min_time_day:
                        errors.append(
                            f"{' '.join(self.user_names[i])} hat am Tag {day+1} nur {int(total_hours / self.hour_devider)} Stunden eingetragen. "
                            f"Das ist weniger als die Mindestarbeitszeit von {int(self.min_time_day / self.hour_devider)} Stunden."
                        )


            if errors:
                raise ValueError("Folgende Fehler wurden gefunden:\n" + "\n".join(errors))
            
            return {"success": True, "name": "Pre-check_6", "message": "All checks are successful!"}
        except ValueError as e:
            return {"success": False, "name": "Pre-check_6", "message": str(e)}


    def pre_check_7(self):
        """
        ---------------------------------------------------------------------------------------------------------------
        7. Ist die min. Zeit pro Tag so klein, dass die Stunden in der gerechten Verteilung nicht erfüllt werden können?
        ---------------------------------------------------------------------------------------------------------------
        """
        return None

    def pre_check_8(self):
        """
        ---------------------------------------------------------------------------------------------------------------
        8. Ist die Toleranz der gerechten Verteilung zu klein gewählt? --> Evtl. die Bedingung weich machen!
        ---------------------------------------------------------------------------------------------------------------
        """
        return None


    def solver_selection(self):
        """
        Auswahl des geeigneten Solvers für Constraint Programmierung.
        """
        self.model = cp_model.CpModel()
        self.solver = cp_model.CpSolver()
        
        self.solver.parameters.max_time_in_seconds = 500 # Der Solver stoppt nach 500s
        self.gap_to_stop = 1                             # Der Solver stoppt unter diesem GAP
        # self.solver.parameters.num_search_workers = 4 # Anzahl Kerne --> noch genau testen was das optimum ist (CPU-Auslastung beachten!)


    def define_penalty_costs(self):
        """
        Definiere Strafkosten für weiche Nebenbedingungen
        """
        # Strafkosten für jede NB
        penalty_values = {
            "nb1": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb2": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb3": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb4": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb5": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb6": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb7": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb8": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb9": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb10": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb11": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000},
            "nb12": {0: 1, 1: 100, 2: 250, 3: 400 , 4: 600, 5: 10000}
        }

        # Mapping für die entsprechenden Namen der Klassenattribute
        penalty_cost_names = {
            "nb1": "penalty_cost_nb1",
            "nb2": "penalty_cost_nb2",
            "nb3": "penalty_cost_nb3_min",
            "nb4": "penalty_cost_nb4_max",
            "nb5": "penalty_cost_nb5_min",
            "nb6": "penalty_cost_nb6_max",
            "nb7": "penalty_cost_nb7",
            "nb8": "penalty_cost_nb8",
            "nb9": "penalty_cost_nb9",
            "nb10": "penalty_cost_nb10",
            "nb11": "penalty_cost_nb11_min",
            "nb12": "penalty_cost_nb12_max"
        }

        # Setze die Strafkosten für jede NB basierend auf dem Dictionary
        for key, values in penalty_values.items():
            if key in self.solver_requirements:
                nb_value = self.solver_requirements[key]
                if nb_value in values:
                    setattr(self, penalty_cost_names[key], values[nb_value])
                    print(f"{penalty_cost_names[key].upper()}: {getattr(self, penalty_cost_names[key])}")
                else:
                    print(f"Zahl für {penalty_cost_names[key].upper()} wurde nicht gefunden")
            else:
                print(f"{penalty_cost_names[key].upper()} nicht in solver_requirements gefunden")



    def decision_variables(self):
        """
        Entscheidungsvariabeln für den CP-Solver
        """
    
        # Arbeitsvariable
        self.x = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[i][j])):
                    woche = j // 7
                    if self.benoetigte_skills[woche]: # Überprüfen ob es in der Woche einen Skilleintrag gibt
                        for s in self.benoetigte_skills[woche]: # Die Skills in der richtigen Woche berücksichtigen
                            if s in self.mitarbeiter_s[i]:  # Dies stellt sicher, dass der Mitarbeiter den Skill tatsächlich hat
                                self.x[i, j, k, s] = self.model.NewIntVar(0, 1, f'x[{i}, {j}, {k}, {s}]')

        # Arbeitsblockvariable
        self.y = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[i][j])):
                    self.y[i, j, k] = self.model.NewIntVar(0, 1, f'y[{i}, {j}, {k}]')

        # Arbeitstagvariable Skillabhängig
        self.a = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for s in self.skills:
                    if s in self.mitarbeiter_s[i]:  # Der Mitarbeiter muss den Skill haben
                        self.a[i, j, s] = self.model.NewBoolVar(f'a[{i}, {j}, {s}]')

        # Arbeitstagvariable Skillunabhängig
        self.a_sum = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.a_sum[i, j] = self.model.NewBoolVar(f'a_sum[{i}, {j}]')


        # Schichtvariable Woche (2-Schicht)
        self.s2 = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.s2[i, j] = self.model.NewIntVar(0, 1, f's2[{i}, {j}]')

        # Schichtvariable Woche (3-Schicht)
        self.s3 = {}
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.s3[i, j] = self.model.NewIntVar(0, 2, f's3[{i}, {j}]')

        # Gleiche Schichten innerhalb 2 und 4 Wochen
        self.c = {}
        for i in self.mitarbeiter:
            for j in range(7, self.calc_time):
                self.c[i, j] = self.model.NewIntVar(0, 1, f'c[{i}, {j}]')


    def violation_variables(self):
        """ 
        Verletzungsvariabeln für den CP-Solver
        """
        
        # Unendlichkeitssimulation -> Soll möglichst vermieden werden!
        INF = int(1e6)

        
        # NB1 violation variable
        for j in range(self.calc_time):
            for k in range(len(self.verfügbarkeit[self.mitarbeiter[0]][j])):
                self.nb1_violation[j, k] = self.model.NewIntVar(0, len(self.mitarbeiter), f'nb1_violation[{j}, {k}]')

        # NB2 violation variable
        diff_1 = self.max_time_week - self.weekly_hours
        # Damit wird sichergestellt, dass diff_1 nicht negativ werden kann.
        if diff_1 < 0:
            diff_1 = 0
        self.nb2_violation = {ma: {} for ma in self.mitarbeiter}
        for ma in self.mitarbeiter:
            for week in range(1, self.week_timeframe + 1):
                self.nb2_violation[ma][week] = self.model.NewIntVar(0, diff_1, f'nb2_violation[{ma}][{week}]')

        # NB3 Mindestarbeitszeit Verletzungsvariable
        diff_2 = self.desired_min_time_day - self.min_time_day
        if diff_2 < 0:
            diff_2 = 0
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.nb3_min_violation[i, j] = self.model.NewIntVar(0, diff_2, f'nb3_min_violation[{i}, {j}]')

        # NB4 Höchstarbeitszeit Verletzungsvariable
        diff_3 = self.max_time_day - self.desired_max_time_day
        if diff_3 < 0:
            diff_3 = 0
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.nb4_max_violation[i, j] = self.model.NewIntVar(0, diff_3, f'nb4_max_violation[{i}, {j}]')

        # NB5 Mindestarbeitszeit Verletzungsvariable
        for i in self.mitarbeiter:
            self.nb5_min_violation[i] = [self.model.NewIntVar(0, self.weekly_hours, f'nb5_min_violation[{i}][{week}]') for week in range(1, self.week_timeframe + 1)]

        # NB6 Höchstarbeitszeit Verletzungsvariable
        for i in self.mitarbeiter:
            self.nb6_max_violation[i] = [self.model.NewIntVar(0, self.weekly_hours, f'nb6_max_violation[{i}][{week}]') for week in range(1, self.week_timeframe + 1)]

        # NB7 Innerhalb einer Woche die gleiche Schicht - Verletzungsvariable
        for i in self.mitarbeiter:
            for j in range(7):
                self.nb7_violation[i, j] = self.model.NewIntVar(0, 1, f'nb7_violation[{i}, {j}]')

        # NB8 Innerhalb der zweiten / vierten Woche die gleiche Schicht - Verletzungsvariable
        for i in self.mitarbeiter:
            for j in range(7, self.calc_time):
                self.nb8_violation[i, j] = self.model.NewIntVar(0, 1, f'nb8_violation[{i}, {j}]')

        # NB9 Minimale Arbeitsstunden pro Block - Verletzungsvariable
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[i][j])):
                    self.nb9_violation[i, j, k] = self.model.NewIntVar(0, 1, f'nb9_violation[{i}, {j}, {k}]')

        # NB10 Max. Anzahl an Arbeitstagen in Folge
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.nb10_violation[i, j] = self.model.NewIntVar(0, 1, f'nb10_violation[{i}, {j}]')

        # NB11 Gerechte Verteilung der Stunden - Unterschreitung
        for i in self.mitarbeiter:
            # 96 * 7 * 4 = 2688
            self.nb11_min_violation[i] = self.model.NewIntVar(0, 2688, f'nb11_min_violation[{i}]')

        # NB12 Gerechte Verteilung der Stunden - Überschreitung
        for i in self.mitarbeiter:
            # 96 * 7 * 4 = 2688
            self.nb12_max_violation[i] = self.model.NewIntVar(0, 2688, f'nb12_max_violation[{i}]')



    def objective_function(self):
        """
        Zielfunktion
        """

        # Kosten MA minimieren
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[i][j])):
                    woche = j // 7
                    if self.benoetigte_skills[woche]:
                        for s in self.benoetigte_skills[woche]:
                            if s in self.mitarbeiter_s[i]:  # Überprüfen, ob der Mitarbeiter den Skill hat
                                self.cost_expressions.append(self.x[i, j, k, s] * int(self.kosten[i] / self.hour_devider))

        # Kosten Weiche NB1
        for j in range(self.calc_time):
            for k in range(len(self.verfügbarkeit[self.mitarbeiter[0]][j])):
                self.cost_expressions.append(self.nb1_violation[j, k] * int(self.penalty_cost_nb1 / self.hour_devider))

        # Kosten Weiche NB2
        for i in self.mitarbeiter:
            for week in range(1, self.week_timeframe + 1):
                self.cost_expressions.append(self.nb2_violation[i][week] * self.penalty_cost_nb2)
        
        # Kosten für Weiche NB3 Mindestarbeitszeit Verletzung
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.cost_expressions.append(self.nb3_min_violation[i, j] * self.penalty_cost_nb3_min)
                
        # Kosten für Weiche NB4 Höchstarbeitszeit Verletzung
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.cost_expressions.append(self.nb4_max_violation[i, j] * self.penalty_cost_nb4_max)

        # Kosten für Weiche NB5 Mindestarbeitszeit Verletzung
        for i in self.mitarbeiter:
            for week in range(1, self.week_timeframe + 1):
                self.cost_expressions.append(self.nb5_min_violation[i][week-1] * self.penalty_cost_nb5_min)

        # Kosten für Weiche NB6 Höchstarbeitszeit Verletzung
        for i in self.mitarbeiter:
            for week in range(1, self.week_timeframe + 1):
                self.cost_expressions.append(self.nb6_max_violation[i][week-1] * self.penalty_cost_nb6_max)
        
        # Kosten für Weiche NB7 "Innerhalb einer Woche immer gleiche Schichten"
        for i in self.mitarbeiter:
            for j in range(7):
                self.cost_expressions.append(self.nb7_violation[i, j] * self.penalty_cost_nb7)
        
        # Kosten für Weiche NB8 "Innerhalb der zweiten Woche immer gleiche Schichten"
        for i in self.mitarbeiter:
            for j in range(7, self.calc_time):
                self.cost_expressions.append(self.nb8_violation[i, j] * self.penalty_cost_nb8)
        
        # Kosten für Weiche NB9 "Minimale Arbeitsstunden pro Block"
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[i][j])):
                    self.cost_expressions.append(self.nb9_violation[i, j, k] * int(self.penalty_cost_nb9 / self.hour_devider))
        
        # Kosten für Weiche NB10 "Max. Anzahl an Arbeitstagen in Folge"
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                self.cost_expressions.append(self.nb10_violation[i, j] * self.penalty_cost_nb10)

        # Kosten für Weiche NB11 "Gerechte Verteilung der Stunden - Unterschreitung"
        for i in self.mitarbeiter:
            self.cost_expressions.append(self.nb11_min_violation[i] * self.penalty_cost_nb11_min)

        # Kosten für Weiche NB12 "Gerechte Verteilung der Stunden - Überschreitung"
        for i in self.mitarbeiter:
            self.cost_expressions.append(self.nb12_max_violation[i] * self.penalty_cost_nb12_max)
        
                
        # Alle Kosten-Ausdrücke summieren und dem Model befehlen, die Kosten zu minimieren
        total_cost = sum(self.cost_expressions)
        self.model.Minimize(total_cost)


    def constraints(self):
        """
        Beschränkungen / Nebenbedingungen
        """

        # -------------------------------------------------------------------------------------------------------
        # HARTE NB -- NEU 08.08.2023 --
        # NB 0 - Variable a ist 1, wenn der Mitarbeiter an einem Tag arbeitet, sonst 0
        # -------------------------------------------------------------------------------------------------------
    
        # 23.10.2023 - a definieren abhängig vom Skill!
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                woche = j // 7
                if self.benoetigte_skills[woche]:
                    for s in self.benoetigte_skills[woche]:
                        if s in self.mitarbeiter_s[i]:  # Prüfen, ob der Mitarbeiter den Skill hat
                            # sum_x ist die Summe, an der ein MA am Tag mit einem bestimmten Skill arbeiten kann
                            sum_x = sum(self.x[i, j, k, s] for k in range(len(self.verfügbarkeit[i][j])))
                            
                            self.model.Add(self.a[i, j, s] * (len(self.verfügbarkeit[i][j]) + 1) >= sum_x)
                            self.model.Add(self.a[i, j, s] <= sum_x)


        # 30.10.2023 - a_sum mit a verknüpfen - Skillunabhängig!
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                # Bsp: sum(self.a[129, 1, s]) = 1 <= 3 * X
                self.model.Add(sum(self.a[i, j, s] for s in self.skills if s in self.mitarbeiter_s[i]) <= len(self.skills) * self.a_sum[i, j])
                # Das verhindert, das bei der summe von 0 a_sum nicht einfach 1 werden darf
                self.model.Add(self.a_sum[i, j] <= sum(self.a[i, j, s] for s in self.skills if s in self.mitarbeiter_s[i]))

    
        # -------------------------------------------------------------------------------------------------------
        # HARTE NB
        # NB 1 - MA nur einteilen, wenn er verfügbar ist. 
        # -------------------------------------------------------------------------------------------------------
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[i][j])):
                    woche = j // 7
                    if self.benoetigte_skills[woche]:
                        for s in self.benoetigte_skills[woche]:
                            if s in self.mitarbeiter_s[i]:  # Prüfen, ob der Mitarbeiter den Skill hat
                                self.model.Add(self.x[i, j, k, s] <= self.verfügbarkeit[i][j][k])



        # -------------------------------------------------------------------------------------------------------
        # HARTE NB
        # NB 2 - Mindestanzahl MA zu jeder Stunde an jedem Tag mit Skill anwesend
        # -------------------------------------------------------------------------------------------------------
        # Stellen Sie sicher, dass zu jeder Stunde eine Mindestanzahl von Mitarbeitern mit dem benötigten Skill anwesend ist
        for j in range(self.calc_time):
            for k in range(len(self.verfügbarkeit[next(iter(self.mitarbeiter_s))][j])):  # Verwendung des ersten Mitarbeiters als Referenz
                for s in self.skills:
                    # Existiert ein Eintrag für die aktuelle Stunde und den aktuellen Skill?
                    if k in self.min_anwesend[s][j]:
                    # if k < len(self.min_anwesend[s][j]) and self.min_anwesend[s][j][k] > 0: Getestet 29.10, bringt auch nichts

                        # Sammeln der Mitarbeiter, die den erforderlichen Skill haben
                        mitarbeiter_mit_skill = [i for i in self.mitarbeiter if s in self.mitarbeiter_s[i]]

                        # Bedingung: Die Anzahl der anwesenden Mitarbeiter muss größer oder gleich der Mindestanzahl sein
                        self.model.Add(sum(self.x[i, j, k, s] for i in mitarbeiter_mit_skill) >= self.min_anwesend[s][j][k])


        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB -- NEU 30.10.2023 --
        # NB 2 - Mindestanzahl MA zu jeder Stunde an jedem Tag mit Skill anwesend 
        # ***** Weiche Nebenbedingung 1 *****
        # -------------------------------------------------------------------------------------------------------
        for j in range(self.calc_time):
            for k in range(len(self.verfügbarkeit[next(iter(self.mitarbeiter_s))][j])):  # Wir nehmen an, dass alle Mitarbeiter die gleichen Öffnungszeiten haben
                for s in self.skills:
                    if k in self.min_anwesend[s][j]:
                        self.model.Add(sum(self.x[i, j, k, s] for i in self.mitarbeiter if s in self.mitarbeiter_s[i]) - self.min_anwesend[s][j][k] <= self.nb1_violation[j, k])


        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB
        # NB 3 - Max. Arbeitszeit pro Woche (Unabhängig vom Skill)
        # ***** Weiche Nebenbedingung 2 *****
        # -------------------------------------------------------------------------------------------------------
        if self.week_timeframe == 1:
            # Berechnung der Gesamtstunden pro Mitarbeiter unter Berücksichtigung der benötigten Skills
            total_hours = {}
            for ma in self.mitarbeiter:
                ma_total_hours = 0  # Initialisieren der Gesamtstunden für den aktuellen Mitarbeiter
                for j in range(self.calc_time):
                    for k in range(len(self.verfügbarkeit[ma][j])):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[ma]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    ma_total_hours += self.x[ma, j, k, s]

                total_hours[ma] = ma_total_hours  # Zuweisung der Gesamtstunden pro Mitarbeiter

            # Hinzufügen der Nebenbedingung für jeden Mitarbeiter
            for ma in self.mitarbeiter:
                self.model.Add(total_hours[ma] - self.weekly_hours <= self.nb2_violation[ma][1])


        elif self.week_timeframe == 2:
            # Berechnung der Gesamtstunden pro Mitarbeiter unter Berücksichtigung der benötigten Skills
            total_hours_week1 = {}
            total_hours_week2 = {}

            # 1. Woche
            for ma in self.mitarbeiter:
                ma_total_hours = 0  # Initialisieren der Gesamtstunden für den aktuellen Mitarbeiter
                for j in range(7): # Für die erste Woche
                    for k in range(len(self.verfügbarkeit[ma][j])):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[ma]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    ma_total_hours += self.x[ma, j, k, s]
                total_hours_week1[ma] = ma_total_hours  # Zuweisung der Gesamtstunden pro Mitarbeiter

            # 2. Woche
            for ma in self.mitarbeiter:
                ma_total_hours = 0  # Initialisieren der Gesamtstunden für den aktuellen Mitarbeiter
                for j in range(7, 14): # Für die zweite Woche
                    for k in range(len(self.verfügbarkeit[ma][j])):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[ma]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    ma_total_hours += self.x[ma, j, k, s]
                total_hours_week2[ma] = ma_total_hours  # Zuweisung der Gesamtstunden pro Mitarbeiter

            # Hinzufügen der Nebenbedingung für jeden Mitarbeiter
            for ma in self.mitarbeiter:
                self.model.Add(total_hours_week1[ma] - self.weekly_hours <= self.nb2_violation[ma][1])
                self.model.Add(total_hours_week2[ma] - self.weekly_hours <= self.nb2_violation[ma][2])

                        
        elif self.week_timeframe == 4:
            # Berechnung der Gesamtstunden pro Mitarbeiter unter Berücksichtigung der benötigten Skills
            total_hours_week1 = {}
            total_hours_week2 = {}
            total_hours_week3 = {}
            total_hours_week4 = {}

            # 1. Woche
            for ma in self.mitarbeiter:
                ma_total_hours = 0  # Initialisieren der Gesamtstunden für den aktuellen Mitarbeiter
                for j in range(7): # Für die erste Woche
                    for k in range(len(self.verfügbarkeit[ma][j])):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[ma]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    ma_total_hours += self.x[ma, j, k, s]
                total_hours_week1[ma] = ma_total_hours  # Zuweisung der Gesamtstunden pro Mitarbeiter

            # 2. Woche
            for ma in self.mitarbeiter:
                ma_total_hours = 0  # Initialisieren der Gesamtstunden für den aktuellen Mitarbeiter
                for j in range(7, 14): # Für die zweite Woche
                    for k in range(len(self.verfügbarkeit[ma][j])):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[ma]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    ma_total_hours += self.x[ma, j, k, s]
                total_hours_week2[ma] = ma_total_hours  # Zuweisung der Gesamtstunden pro Mitarbeiter

            # 3. Woche
            for ma in self.mitarbeiter:
                ma_total_hours = 0  # Initialisieren der Gesamtstunden für den aktuellen Mitarbeiter
                for j in range(14, 21): # Für die dritte Woche
                    for k in range(len(self.verfügbarkeit[ma][j])):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[ma]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    ma_total_hours += self.x[ma, j, k, s]
                total_hours_week3[ma] = ma_total_hours  # Zuweisung der Gesamtstunden pro Mitarbeiter

            # 4. Woche
            for ma in self.mitarbeiter:
                ma_total_hours = 0  # Initialisieren der Gesamtstunden für den aktuellen Mitarbeiter
                for j in range(21, 28): # Für die vierte Woche
                    for k in range(len(self.verfügbarkeit[ma][j])):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[ma]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    ma_total_hours += self.x[ma, j, k, s]
                total_hours_week4[ma] = ma_total_hours  # Zuweisung der Gesamtstunden pro Mitarbeiter

            # Hinzufügen der Nebenbedingung für jeden Mitarbeiter
            for ma in self.mitarbeiter:
                self.model.Add(total_hours_week1[ma] - self.weekly_hours <= self.nb2_violation[ma][1])
                self.model.Add(total_hours_week2[ma] - self.weekly_hours <= self.nb2_violation[ma][2])
                self.model.Add(total_hours_week3[ma] - self.weekly_hours <= self.nb2_violation[ma][3])
                self.model.Add(total_hours_week4[ma] - self.weekly_hours <= self.nb2_violation[ma][4])


        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB -- NEU 31.07.2023 --
        # NB 4 - Min. und Max. Arbeitszeit pro Tag (Skillunabhängig)
        # ***** Weiche Nebenbedingung 3 und 4 *****
        # -------------------------------------------------------------------------------------------------------
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                woche = j // 7
                if self.benoetigte_skills[woche]:
                    # Summieren Sie die Stunden über alle benötigten Skills pro Tag, die der Mitarbeiter besitzt
                    sum_hour = sum(self.x[i, j, k, s] for s in self.benoetigte_skills[woche] 
                                for k in range(len(self.verfügbarkeit[i][j])) 
                                if s in self.mitarbeiter_s[i])
                    
                    # Prüfen, ob die Summe der Arbeitsstunden kleiner als die Mindestarbeitszeit ist
                    self.model.Add(sum_hour - self.min_zeit[i] * self.a_sum[i, j] >= -self.nb3_min_violation[i, j])
                    self.model.Add(self.nb3_min_violation[i, j] >= 0)

                    # Prüfen, ob die Summe der Arbeitsstunden größer als die maximale Arbeitszeit ist
                    self.model.Add(sum_hour - self.max_zeit[i] * self.a_sum[i, j] <= self.nb4_max_violation[i, j])
                    self.model.Add(self.nb4_max_violation[i, j] >= 0)
        

        # -------------------------------------------------------------------------------------------------------
        # HARTE NB
        # NB 5 - Anzahl Arbeitsblöcke
        # -------------------------------------------------------------------------------------------------------
        """
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                # Überprüfen, ob der Betrieb an diesem Tag geöffnet ist
                if self.opening_hours[j] > 0:
                    # Für die erste Stunde des Tages
                    self.model.Add(self.y[i, j, 0] >= self.x[i, j, 0])
                    
                    # Für die restlichen Stunden des Tages
                    for k in range(1, len(self.verfügbarkeit[i][j])):
                        self.model.Add(self.y[i, j, k] >= self.x[i, j, k] - self.x[i, j, k-1])

                    # Die Summe der y[i, j, k] für einen bestimmten Tag j sollte nicht größer als daily_deployment sein
                    self.model.Add(sum(self.y[i, j, k] for k in range(len(self.verfügbarkeit[i][j]))) <= self.daily_deployment)
        """

        # 26.10.2023 (01.11.2023: NB sollte eig funktionieren)
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                # Überprüfen, ob der Betrieb an diesem Tag geöffnet ist
                if self.opening_hours[j] > 0:
                    current_week = j // 7

                    # Benötigten Skills für die aktuelle Woche
                    needed_skills_this_week = self.benoetigte_skills[current_week]

                    # Filtern der Skills, um nur diejenigen zu behalten, die der Mitarbeiter für den j-ten Tag hat
                    valid_skills = [s for s in needed_skills_this_week if s in self.mitarbeiter_s[i]]

                    # Für die erste Stunde des Tages berechnen wir die Summe unter Berücksichtigung der gültigen Skills
                    self.model.Add(self.y[i, j, 0] >= sum(self.x[i, j, 0, s] for s in valid_skills))
                    
                    # Für die restlichen Stunden des Tages berechnen wir die Summe unter Berücksichtigung der gültigen Skills
                    for k in range(1, len(self.verfügbarkeit[i][j])):
                        self.model.Add(self.y[i, j, k] >= sum(self.x[i, j, k, s] for s in valid_skills) - sum(self.x[i, j, k-1, s] for s in valid_skills))


                    # Die Summe der y[i, j, k] für einen bestimmten Tag j sollte nicht größer als daily_deployment sein
                    self.model.Add(sum(self.y[i, j, k] for k in range(len(self.verfügbarkeit[i][j]))) <= self.daily_deployment)


        
        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB -- NEU 08.08.23 --
        # NB 6 - "Perm" Mitarbeiter zu employement_level fest einplanen
        # ***** Weiche Nebenbedingung 5 und 6 *****
        # -------------------------------------------------------------------------------------------------------
        """
        for i, ma in enumerate(self.mitarbeiter):
            if self.employment[i] == "Perm":
                for week in range(1, self.week_timeframe + 1):
                    week_start = (week - 1) * (self.calc_time // self.week_timeframe)
                    week_end = week * (self.calc_time // self.week_timeframe)
                    # print(f'Week start for employee {ma} in week {week}: {week_start}')
                    # print(f'Week end for employee {ma} in week {week}: {week_end}')

                    total_hours_week = sum(self.x[ma, j, k] for j in range(week_start, week_end) for k in range(len(self.verfügbarkeit[ma][j])))
                    # print(f'Total hours for employee {ma} in week {week}: {total_hours_week}')

                    # Prüfen, ob die Gesamtstunden kleiner als die vorgegebenen Arbeitsstunden sind (Unterschreitung)
                    # Die Wochenstunden werden mit dem employment_lvl multipliziert und aufgerundet!
                    self.model.Add(total_hours_week - round((self.weekly_hours * self.employment_lvl_exact[i]) + 0.5) >= -self.nb5_min_violation[ma][week - 1])
                    self.model.Add(self.nb5_min_violation[ma][week - 1] >= 0)

                    # Prüfen, ob die Gesamtstunden größer als die vorgegebenen Arbeitsstunden sind (Überschreitung)
                    self.model.Add(round((self.weekly_hours * self.employment_lvl_exact[i]) + 0.5) - total_hours_week >= -self.nb6_max_violation[ma][week - 1])
                    self.model.Add(self.nb6_max_violation[ma][week - 1] >= 0)
        """

        # 23.10.2023
        for i, ma in enumerate(self.mitarbeiter):
            if self.employment[i] == "Perm":
                for week in range(1, self.week_timeframe + 1):
                    week_start = (week - 1) * (self.calc_time // self.week_timeframe)
                    week_end = week * (self.calc_time // self.week_timeframe)

                    # Ermitteln der benötigten Skills für die spezifische Woche
                    needed_skills_this_week = []
                    if week - 1 < len(self.benoetigte_skills):  # Sicherstellen, dass wir nicht über die Liste hinausgehen
                        needed_skills_this_week = self.benoetigte_skills[week - 1]

                    # Summieren über alle gültigen Skills, die der Mitarbeiter ma hat und die für diese Woche benötigt werden
                    total_hours_week = sum(
                        self.x[ma, j, k, s]
                        for j in range(week_start, week_end)
                        for k in range(len(self.verfügbarkeit[ma][j]))
                        for s in needed_skills_this_week  # Nutze nur die Skills, die in dieser Woche benötigt werden
                        if s in self.mitarbeiter_s[ma] and j < len(self.verfügbarkeit[ma])  # Sicherstellen, dass der Index gültig ist
                    )

                    # Unterschreitung
                    min_hours = round((self.weekly_hours * self.employment_lvl_exact[i]) + 0.5)
                    self.model.Add(total_hours_week - min_hours >= -self.nb5_min_violation[ma][week - 1])
                    self.model.Add(self.nb5_min_violation[ma][week - 1] >= 0)

                    # Überschreitung
                    max_hours = round((self.weekly_hours * self.employment_lvl_exact[i]) + 0.5)
                    self.model.Add(max_hours - total_hours_week >= -self.nb6_max_violation[ma][week - 1])
                    self.model.Add(self.nb6_max_violation[ma][week - 1] >= 0)



        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB
        # NB 7 - Innerhalb einer Woche immer gleiche Schichten
        # ***** Weiche Nebenbedingung 7 *****
        # -------------------------------------------------------------------------------------------------------
        if self.company_shifts <= 1:
            pass

        elif self.company_shifts == 2:
            for i in self.mitarbeiter:
                for j in range(7):

                    # Bei z.B. 7 Stunden bekommt die erste Schicht 3 Stunden und die zweite Schicht 4 Stunden!
                    first_shift_hours = 0
                    second_shift_hours = 0

                    # Über alle Stunden in der ersten Schichthälfte summieren
                    for k in range(int(len(self.verfügbarkeit[i][j]) / 2)):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[i]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    first_shift_hours += self.x[i, j, k, s]

                    # Über alle Stunden in der zweiten Schichthälfte summieren
                    for k in range(int(len(self.verfügbarkeit[i][j]) / 2), len(self.verfügbarkeit[i][j])):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[i]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    second_shift_hours += self.x[i, j, k, s]
                    
                    delta = self.model.NewBoolVar("delta")
                    
                    self.model.Add(first_shift_hours - second_shift_hours - 1000 * delta <= 0)         
                    self.model.Add(second_shift_hours - first_shift_hours - 1000 * (1 - delta) <= 0)   
                    
                    self.model.Add(self.s2[i, j] == 1 - delta)

            """
            # Harte nb Option zum testen
            for i in self.mitarbeiter:
                for j in range(1, self.calc_time):
                    self.model.Add(self.s2[i, j] - self.s2[i, j-1] == 0)
            """ 

            # Bedingungen, um sicherzustellen, dass innerhalb einer Woche immer die gleiche Schicht gearbeitet wird
            for i in self.mitarbeiter:
                for j in range(1, 7):
                    diff = self.model.NewIntVar(-1, 1, "diff")
                    
                    self.model.Add(diff == self.s2[i, j] - self.s2[i, j-1])
            
                    # Bedingungen für den "absoluten Wert"
                    self.model.Add(self.nb7_violation[i, j] >= diff)
                    self.model.Add(self.nb7_violation[i, j] >= -diff)



        elif self.company_shifts == 3:
            for i in self.mitarbeiter:
                for j in range(7):
                    total_len = len(self.verfügbarkeit[i][j])
                    third_shift_len = total_len // 3
                    first_shift_len = third_shift_len + (total_len % 3) // 2
                    second_shift_len = third_shift_len + (total_len % 3) - (total_len % 3) // 2

                    first_shift_hours = 0
                    second_shift_hours = 0
                    third_shift_hours = 0

                    # Erst Schicht des Tages
                    for k in range(0, first_shift_len):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[i]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    first_shift_hours += self.x[i, j, k, s]

                    # Zweite Schicht des Tages
                    for k in range(first_shift_len, first_shift_len + second_shift_len):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[i]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    second_shift_hours += self.x[i, j, k, s]
                    
                    # Dritte Schicht des Tages
                    for k in range(first_shift_len + second_shift_len, total_len):
                        if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                            for s in self.benoetigte_skills[0]:
                                if s in self.mitarbeiter_s[i]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                    third_shift_hours += self.x[i, j, k, s]

                    delta1 = self.model.NewBoolVar(f"delta1_{i}_{j}")
                    delta2 = self.model.NewBoolVar(f"delta2_{i}_{j}")
                    delta3 = self.model.NewBoolVar(f"delta3_{i}_{j}")

                    M = 1000

                    self.model.Add(first_shift_hours >= second_shift_hours - M * (1 - delta1))
                    self.model.Add(first_shift_hours >= third_shift_hours - M * (1 - delta1))

                    self.model.Add(second_shift_hours >= first_shift_hours - M * (1 - delta2))
                    self.model.Add(second_shift_hours >= third_shift_hours - M * (1 - delta2))

                    self.model.Add(third_shift_hours >= first_shift_hours - M * (1 - delta3))
                    self.model.Add(third_shift_hours >= second_shift_hours - M * (1 - delta3))

                    self.model.Add(delta1 + delta2 + delta3 == 1)

                    self.model.Add(self.s3[i, j] == 0 * delta1 + 1 * delta2 + 2 * delta3)

            """
            # Harte Bedingung, dass innerhalb einer Woche immer die gleiche Schicht gearbeitet wird
            for i in self.mitarbeiter:
                for j in range(1, self.calc_time):
                    self.model.Add(self.s3[i, j] - self.s3[i, j-1] == 0)
            """

            # Bedingungen, um sicherzustellen, dass innerhalb einer Woche immer die gleiche Schicht gearbeitet wird
            for i in self.mitarbeiter:
                for j in range(1, 7):
                    diff = self.model.NewIntVar(-2, 2, "diff") # Unterschied kann -2, -1, 0, 1 oder 2 sein
                    self.model.Add(diff == self.s3[i, j] - self.s3[i, j-1])

                    self.model.Add(self.nb7_violation[i, j] >= diff)
                    self.model.Add(self.nb7_violation[i, j] >= -diff)
                    self.model.Add(self.nb7_violation[i, j] <= 1) # Die Verletzung sollte maximal 1 betragen

        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB (04.11.2023)
        # NB 8 - Wechselnde Schichten innerhalb von 2 und 4 Wochen
        # ***** Weiche Nebenbedingung 8 *****
        # -------------------------------------------------------------------------------------------------------

        # 2 Wochen + 2-Schicht # --------------------------------------------------------------------------------
        if self.week_timeframe == 2:
            if self.company_shifts == 2:
                for i in self.mitarbeiter:
                    # Anzahl der Tage in der ersten Woche, an denen in der ersten bzw. zweiten Schicht gearbeitet wurde
                    first_week_first_shift_days = sum(self.s2[i, j] for j in range(7))
                    first_week_second_shift_days = 7 - first_week_first_shift_days

                    # Hilfsvariable, um die Schicht der ersten Woche festzulegen
                    first_week_shift = self.model.NewBoolVar("first_week_shift")

                    # Wenn die Anzahl der Tage in der ersten Schicht größer ist, setzen Sie first_week_shift auf 1
                    self.model.Add(first_week_first_shift_days - first_week_second_shift_days - 1000 * first_week_shift <= 0)
                    self.model.Add(first_week_second_shift_days - first_week_first_shift_days - 1000 * (1 - first_week_shift) <= 0)

                    for j in range(7, 14):
                        self.model.Add(self.c[i, j] == 1 - first_week_shift)

                    # In der zweiten Woche muss der Mitarbeiter in der entgegengesetzten Schicht arbeiten
                    for j in range(7, 14):
                        # Summe der Stunden in der ersten Schicht in der zweiten Woche
                        first_shift_hours_second_week = 0
                        second_shift_hours_second_week = 0

                        for k in range(0, int(len(self.verfügbarkeit[i][j]) / 2)):
                            if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                                for s in self.benoetigte_skills[0]:
                                    if s in self.mitarbeiter_s[i]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                        first_shift_hours_second_week += self.x[i, j, k, s]

                        for k in range(int(len(self.verfügbarkeit[i][j]) / 2), len(self.verfügbarkeit[i][j])):
                            if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                                for s in self.benoetigte_skills[0]:
                                    if s in self.mitarbeiter_s[i]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                        second_shift_hours_second_week += self.x[i, j, k, s]

                        # first_shift_hours_second_week = sum(self.x[i, j, k] for k in range(0, int(len(self.verfügbarkeit[i][j]) / 2)))
                        # second_shift_hours_second_week = sum(self.x[i, j, k] for k in range(int(len(self.verfügbarkeit[i][j]) / 2), len(self.verfügbarkeit[i][j])))

                        # Kann 0 oder 1 annehmen
                        delta_2 = self.model.NewBoolVar("delta_2")
                        
                        self.model.Add(first_shift_hours_second_week - second_shift_hours_second_week - 1000 * delta_2 <= 0)      
                        self.model.Add(second_shift_hours_second_week - first_shift_hours_second_week - 1000 * (1 - delta_2) <= 0)  
                        
                        # Hilfsvariable mit s2[i, j] verknüpfen
                        self.model.Add(self.s2[i, j] == 1 - delta_2)
                        
                        # Harte Nebenbedingung
                        # self.model.Add(self.s2[i, j] == self.c[i, j])

                # Verletzungsvariable erhöhen, wenn die Schicht in der zweiten Woche nicht der entgegengesetzten Schicht entspricht
                for i in self.mitarbeiter:
                    for j in range(7, 14):
                        diff = self.model.NewIntVar(-1, 1, f"diff_{i}_{j}")
                        self.model.Add(diff == self.s2[i, j] - self.c[i, j])

                        self.model.Add(self.nb8_violation[i, j] >= diff)
                        self.model.Add(self.nb8_violation[i, j] >= -diff)


        # 2 Wochen + 3-Schicht # --------------------------------------------------------------------------------

        """
        Der Code muss neu erstellt werden, viel zu gross in or_algorithm und funktionierte auch noch nicht richtig
        """

        # 4 Wochen + 2-Schicht ----------------------------------------------------------------------------------
        if self.week_timeframe == 4:
            if self.company_shifts == 2:
                for i in self.mitarbeiter:
                    # Anzahl der Tage in der ersten Woche, an denen in der ersten bzw. zweiten Schicht gearbeitet wurde
                    first_week_first_shift_days = sum(self.s2[i, j] for j in range(7))
                    first_week_second_shift_days = 7 - first_week_first_shift_days

                    # Hilfsvariable, um die Schicht der ersten Woche festzulegen
                    first_week_shift = self.model.NewBoolVar("first_week_shift")

                    # Wenn die Anzahl der Tage in der ersten Schicht größer ist, setzen Sie first_week_shift auf 1
                    self.model.Add(first_week_first_shift_days - first_week_second_shift_days - 1000 * first_week_shift <= 0)
                    self.model.Add(first_week_second_shift_days - first_week_first_shift_days - 1000 * (1 - first_week_shift) <= 0)

                    for week in range(2, 5): # Beginne bei Woche 2
                        for j in range((week-1)*7, week*7): # Beginne bei Tag 0 der aktuellen Woche und gehe bis zum letzten Tag der Woche
                            if week % 2 == 0:
                                self.model.Add(self.c[i, j] == 1 - first_week_shift)
                            else:
                                self.model.Add(self.c[i, j] == first_week_shift)

                    for week in range(2, 5): # Wochen 2 bis 4
                        for j in range((week-1)*7, week*7):
                            # Summe der Stunden in der ersten Schicht in der aktuellen Woche
                            first_shift_hours_current_week = 0
                            second_shift_hours_current_week = 0

                            for k in range(0, int(len(self.verfügbarkeit[i][j]) / 2)):
                                if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                                    for s in self.benoetigte_skills[0]:
                                        if s in self.mitarbeiter_s[i]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                            first_shift_hours_current_week += self.x[i, j, k, s]

                            for k in range(int(len(self.verfügbarkeit[i][j]) / 2), len(self.verfügbarkeit[i][j])):
                                if self.benoetigte_skills[0]:  # Überprüfen, ob in der Woche Skills benötigt werden
                                    for s in self.benoetigte_skills[0]:
                                        if s in self.mitarbeiter_s[i]:  # Sicherstellen, dass der Mitarbeiter den Skill hat
                                            second_shift_hours_current_week += self.x[i, j, k, s]

                                        
                            # first_shift_hours_current_week = sum(self.x[i, j, k] for k in range(0, int(len(self.verfügbarkeit[i][j]) / 2)))
                            # second_shift_hours_current_week = sum(self.x[i, j, k] for k in range(int(len(self.verfügbarkeit[i][j]) / 2), len(self.verfügbarkeit[i][j])))

                            # Kann 0 oder 1 annehmen
                            delta_2 = self.model.NewBoolVar("delta_2")
                            
                            self.model.Add(first_shift_hours_current_week - second_shift_hours_current_week - 1000 * delta_2 <= 0)      
                            self.model.Add(second_shift_hours_current_week - first_shift_hours_current_week - 1000 * (1 - delta_2) <= 0)  
                            
                            # Hilfsvariable mit s2[i, j] verknüpfen
                            self.model.Add(self.s2[i, j] == 1 - delta_2)
                            
                            # Harte Nebenbedingung
                            # self.model.Add(self.s2[i, j] == self.c[i, j])

                # Verletzungsvariable erhöhen, wenn die Schicht in den Wochen 2, 3 und 4 nicht der entgegengesetzten Schicht entspricht
                for i in self.mitarbeiter:
                    for j in range(7, 28):
                        diff = self.model.NewIntVar(-1, 1, f"diff_{i}_{j}")
                        self.model.Add(diff == self.s2[i, j] - self.c[i, j])

                        self.model.Add(self.nb8_violation[i, j] >= diff)
                        self.model.Add(self.nb8_violation[i, j] >= -diff)


        # 4 Wochen + 3-Schicht ----------------------------------------------------------------------------------
        if self.week_timeframe == 4:
            if self.company_shifts == 3:
                for i in self.mitarbeiter:
                    pass


        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB (02.11.2023)
        # NB 9 - Minimale Arbeitsstunden pro Arbeitsblock
        # ***** Weiche Nebenbedingung 9 *****
        # -------------------------------------------------------------------------------------------------------
        self.time_per_deployment = self.time_per_deployment * self.hour_devider

        if self.daily_deployment == 2:
            for i in self.mitarbeiter:  
                for j in range(self.calc_time):
                    for k in range(len(self.verfügbarkeit[i][j])):
                        # Bestimmen der gültigen Skills für den Mitarbeiter zu diesem Zeitpunkt
                        current_week = j // 7
                        valid_skills = [s for s in self.benoetigte_skills[current_week] if s in self.mitarbeiter_s[i]]
                        print(valid_skills)
                        
                        # Für alle Stunden minus die mindestanzahl Stunden
                        if k < len(self.verfügbarkeit[i][j]) - self.time_per_deployment + 1:
                            working_conditions = [self.y[i, j, k] - sum(self.x[i, j, k, s] for s in valid_skills)]
                            for h in range(1, self.time_per_deployment):
                                if k + h < len(self.verfügbarkeit[i][j]):
                                    working_conditions.append(self.y[i, j, k] - sum(self.x[i, j, k + h, s] for s in valid_skills))

                            self.model.Add(sum(working_conditions) <= self.nb9_violation[i, j, k])

                        # Für die letzten Stunden des Tages
                        else:
                            # Verbleibende Stunden des Tages ermitteln
                            remaining_hours = len(self.verfügbarkeit[i][j]) - k
                            # Anzahl der Stunden ermitteln, die fehlen, um self.time_per_deployment zu erreichen
                            missing_hours = self.time_per_deployment - remaining_hours

                            # Anzahl der Verstöße ist gleich der Anzahl der fehlenden Stunden
                            self.model.Add(self.y[i, j, k] * missing_hours == self.nb9_violation[i, j, k])

        
        # -------------------------------------------------------------------------------------------------------
        # HARTE NB (08.10.2023)
        # NB 10 - Max. Anzahl an Arbeitstagen in Folge -> Obergrenze
        # -------------------------------------------------------------------------------------------------------
    
        # Diese Variable noch in der Datenbank implementieren
        self.subsequent_workingdays_max = 4

        for i in self.mitarbeiter:
            for j in range(self.calc_time - self.subsequent_workingdays_max):
                # Rollfenster-Technik:
                consecutive_days = [self.a_sum[i, j + k] for k in range(self.subsequent_workingdays_max + 1)]

                # Lineares Ausdruck-Objekt für die Summe der aufeinanderfolgenden Tage
                sum_consecutive_days = sum(consecutive_days)
                
                # Summe der Arbeitstage soll nicht größer als self.subsequent_workingdays_max sein
                self.model.Add(sum_consecutive_days <= self.subsequent_workingdays_max)
        # -------------------------------------------------------------------------------------------------------
        # WEICHE NB (30.08.2023)
        # NB 11 - Max. Anzahl an Arbeitstagen in Folge
        # ***** Weiche Nebenbedingung 10 *****
        # -------------------------------------------------------------------------------------------------------
        for i in self.mitarbeiter:
            for j in range(self.calc_time - self.subsequent_workingdays):
                # Rollfenster-Technik:
                consecutive_days = [self.a_sum[i, j + k] for k in range(self.subsequent_workingdays + 1)]
                
                # Lineares Ausdruck-Objekt für die Summe der aufeinanderfolgenden Tage
                sum_consecutive_days = sum(consecutive_days)
                
                # Summe der Arbeitstage soll nicht größer als self.subsequent_workingdays sein
                self.model.Add(sum_consecutive_days <= self.subsequent_workingdays + self.nb10_violation[i, j])
                # nb10_violation wird nur erhöht, wenn die Anzahl der aufeinanderfolgenden Arbeitstage das Limit überschreitet
                self.model.Add(self.nb10_violation[i, j] >= sum_consecutive_days - self.subsequent_workingdays)
                self.model.Add(self.nb10_violation[i, j] >= 0)
        
        

        # -------------------------------------------------------------------------------------------------------
        # Weiche NB (08.10.2023)
        # NB 12 - Verteilungsgrad MA Über- und Unterschreitung
        # ***** Weiche Nebenbedingung 11 und 12 *****
        # -------------------------------------------------------------------------------------------------------
        
        # Harte NB (mit self.fair_distribution als Toleranz)
        """
        verteilungsstunden = {ma: sum(self.x[ma, j, k] for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[ma][j]))) for ma in self.mitarbeiter}

        for i, ma in enumerate(self.mitarbeiter):
            lower_bound = math.floor(self.gerechte_verteilung[i] * (1 - self.fair_distribution))
            upper_bound = math.ceil(self.gerechte_verteilung[i] * (1 + self.fair_distribution))
            
            self.model.Add(verteilungsstunden[ma] <= upper_bound)
            self.model.Add(verteilungsstunden[ma] >= lower_bound)
        """

        # ACHTUNG: Die Penalty-Costs für eine Verletzung sollten tiefer als der Rest gewählt werden!
        
        # 26.10.2023
        # Initialisierung eines Dictionaries zur Speicherung der Gesamtstunden pro Mitarbeiter
        gesamtstunden_pro_ma = {ma: 0 for ma in self.mitarbeiter}

        # Berechnung der Gesamtstunden für jeden Mitarbeiter, unabhängig vom Skill
        for ma in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[ma][j])):
                    woche = j // 7
                    if self.benoetigte_skills[woche]:  # Überprüfen, ob in der Woche Skills benötigt werden
                        for s in self.benoetigte_skills[woche]:  # Die Skills in der richtigen Woche berücksichtigen
                            if s in self.mitarbeiter_s[ma]:  # Stellen Sie sicher, dass der Mitarbeiter den Skill hat
                                # Verwenden Sie die direkte Key-Zugriffsmethode, da wir sicher sind, dass sie existiert
                                stunden = self.x[ma, j, k, s] if (ma, j, k, s) in self.x else 0
                                gesamtstunden_pro_ma[ma] += stunden

        # Hinzufügen von Bedingungen an das Modell für jeden Mitarbeiter
        for i, ma in enumerate(self.mitarbeiter):
            # Vergleich der gesamten Arbeitsstunden des Mitarbeiters mit der gerechten Verteilung
            self.model.Add(gesamtstunden_pro_ma[ma] >= self.gerechte_verteilung[i] - self.nb11_min_violation[ma])
            self.model.Add(gesamtstunden_pro_ma[ma] <= self.gerechte_verteilung[i] + self.nb12_max_violation[ma])

            # Die Verletzungsvariablen sollten immer nicht-negativ sein
            self.model.Add(self.nb11_min_violation[ma] >= 0)
            self.model.Add(self.nb12_max_violation[ma] >= 0)
       

        # -------------------------------------------------------------------------------------------------------
        # HARTE NB (01.11.2023)
        # NB 13 - Ein Mitarbeiter darf nie mit verschiedenen Skills in der gleichen Stunde arbeiten
        # -------------------------------------------------------------------------------------------------------
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[i][j])):
                    # Prüfen, ob Mitarbeiter in dieser Stunde für mehr als einen Skill verfügbar ist
                    if len([s for s in self.mitarbeiter_s[i] if (i, j, k, s) in self.x]) > 1:
                        # Die Summe der zugewiesenen Skills für diesen Mitarbeiter in dieser Stunde sollte <= 1 sein
                        self.model.Add(sum(self.x[i, j, k, s] for s in self.mitarbeiter_s[i]) <= 1)
        

        # -------------------------------------------------------------------------------------------------------
        # HARTE NB (02.11.2023)
        # NB 14 - Mehrere Skills an einem Tag ausführen J/N
        # -------------------------------------------------------------------------------------------------------
        # Abfragen ob bei mehreren Skills die Mitarbeiter innerhalb von einem Tag die Skills wechseln dürfen.
        # Wenn nein, wird diese NB aktiviert
        # Das muss noch in die Datenbank eingebaut werden
        self.skills_per_day = 1

        if self.skills_per_day == 1:
            for i in self.mitarbeiter:
                for j in range(self.calc_time):
                    # Sicherstellen, dass die Summe der a[i, j, s] über alle Skills s an einem Tag j für jeden Mitarbeiter i kleiner oder gleich 1 ist.
                    self.model.Add(sum(self.a[i, j, s] for s in self.mitarbeiter_s[i] if j // 7 == woche and s in self.benoetigte_skills[woche]) <= 1)




    def solve_problem(self):
        """
        Problem lösen und Kosten ausgeben
        """
        self.last_best_value = None
        self.current_best_value = None
        self.last_read_time = 0
        self.stop_thread = False

        # -----------------------------
        # Nach wievielen Sekunden soll jeweils der aktuelle Bestwert in der Liste aufgenommen werden:
        self.best_test_time = 5
        # -----------------------------

        start_time_solver = time.time()

        log_file = open('log.txt', 'w')

        def log_callback(message):
            log_file.write(message)
            log_file.flush()
            self.read_best_value(message)
            current_time = time.time()
            if current_time - self.last_read_time > self.best_test_time:
                self.last_read_time = current_time

        def add_best_value_to_list():
            hiring_cost_min = self.verteilbare_stunden / self.hour_devider * 100
            while not self.stop_thread:
                time.sleep(self.best_test_time)
                if self.last_best_value is not None:
                    self.current_best_value = self.last_best_value
                self.best_values.append(self.current_best_value)
                gap_now = round(((1 - (hiring_cost_min / self.current_best_value))*100), 2)
                print("------------------ NEUER WERT HINZUGEFÜGT: ", self.current_best_value, "GAP:", gap_now, "%", "------------------------")

                # Überprüfen, ob der Gap weniger als 1% beträgt, wenn ja, solver stoppen
                if gap_now < self.gap_to_stop:
                    print("Gap unter 1%, stoppe den Solver.")
                    self.solver.StopSearch()
                    self.stop_thread = True  # Thread in diesem Fall sofort stoppen.


        self.solver.parameters.log_search_progress = True
        self.solver.parameters.log_to_stdout = True
        self.solver.log_callback = log_callback

        thread = threading.Thread(target=add_best_value_to_list)
        thread.start()

        self.status = self.solver.Solve(self.model)

        end_time_solver = time.time()
        self.solving_time_seconds = end_time_solver - start_time_solver

        self.stop_thread = True
        thread.join()
        log_file.close()

        print("Best Values: ", self.best_values)

    def read_best_value(self, message):
        match = re.search(r'best:(\d+)', message)
        if match:
            self.last_best_value = int(match.group(1))



    def calculate_costs(self):
        """
        In dieser Methode werden die Kosten der einzelnen Nebenbedingungen berechnet und ausgegeben
        """
        self.model.ExportToFile('slow_model.pb.txt')

        
        # ----------------------------------------------------------------
        # Die Werte von s2 printen
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                # Drucken Sie den Wert von s2[i, j]
                print(f"s2[{i}][{j}] =", self.solver.Value(self.s2[i, j]))
        

        # Die Werte von a printen
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for s in self.skills:
                    if s in self.mitarbeiter_s[i]:
                        # Drucken Sie den Wert von a[i, j]
                        print(f"a[{i}][{j}][{s}] =", self.solver.Value(self.a[i, j, s]))

        # Arbeitstagvariable Skillunabhängig
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                print(f"a_sum[{i}][{j}] =", self.solver.Value(self.a_sum[i, j]))
        
        # ----------------------------------------------------------------

        """
        # Die Werte von x ausgeben, wenn sie existieren
        for i in self.mitarbeiter:
            for j in range(self.calc_time):
                for k in range(len(self.verfügbarkeit[i][j])):
                    # Beachten Sie, dass Sie hier über self.benoetigte_skills gehen, um nur die relevanten Skills zu prüfen
                    woche = j // 7
                    if self.benoetigte_skills[woche]:  # Überprüfen, ob in der Woche Skills benötigt werden
                        for s in self.benoetigte_skills[woche]:  # Die Skills in der richtigen Woche berücksichtigen
                            if s in self.mitarbeiter_s[i]:  # Dies stellt sicher, dass der Mitarbeiter den Skill tatsächlich hat
                                # Prüfen, ob der Key existiert, bevor auf das Element zugegriffen wird
                                key = (i, j, k, s)
                                if key in self.x:
                                    print(f"x[{i}][{j}][{k}][{s}] =", self.solver.Value(self.x[key]))
        """



        # Kosten für die Einstellung von Mitarbeitern
        # self.hiring_costs = sum((self.kosten[i] / self.hour_devider) * self.solver.Value(self.x[i, j, k]) for i in self.mitarbeiter for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[i][j])))

        # 23.10.2023
        # Kosten für die Einstellung von Mitarbeitern
        # Kosten für die Einstellung von Mitarbeitern berechnen
        self.hiring_costs = sum(
            (self.kosten[i] / self.hour_devider) * self.solver.Value(self.x.get((i, j, k, s), 0))
            for i in self.mitarbeiter
            for j in range(self.calc_time)
            for k in range(len(self.verfügbarkeit[i][j]))
            for s in [skill for skill in self.skills if skill in self.mitarbeiter_s[i]]
            if (i, j, k, s) in self.x  # Diese Bedingung stellt sicher, dass der Schlüssel existiert
        )

        
        self.violation_nb1 = sum(self.solver.Value(self.nb1_violation[j, k]) for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[self.mitarbeiter[0]][j])))
        self.nb1_penalty_costs = (self.penalty_cost_nb1 / self.hour_devider) * self.violation_nb1

        self.violation_nb2 = sum(self.solver.Value(self.nb2_violation[i][week]) for i in self.mitarbeiter for week in range(1, self.week_timeframe + 1))
        self.nb2_penalty_costs = self.penalty_cost_nb2 * self.violation_nb2

        self.violation_nb3 = sum(self.solver.Value(self.nb3_min_violation[i, j]) for i in self.mitarbeiter for j in range(self.calc_time))
        self.nb3_min_penalty_costs = self.penalty_cost_nb3_min * self.violation_nb3

        self.violation_nb4 = sum(self.solver.Value(self.nb4_max_violation[i, j]) for i in self.mitarbeiter for j in range(self.calc_time))
        self.nb4_max_penalty_costs = self.penalty_cost_nb4_max * self.violation_nb4

        self.violation_nb5 = sum(self.solver.Value(self.nb5_min_violation[i][week - 1]) for i in self.mitarbeiter for week in range(1, self.week_timeframe + 1))
        self.nb5_min_penalty_costs = self.penalty_cost_nb5_min * self.violation_nb5

        self.violation_nb6 = sum(self.solver.Value(self.nb6_max_violation[i][week - 1]) for i in self.mitarbeiter for week in range(1, self.week_timeframe + 1))
        self.nb6_max_penalty_costs = self.penalty_cost_nb6_max * self.violation_nb6

        self.violation_nb7 = sum(self.solver.Value(self.nb7_violation[i, j]) for i in self.mitarbeiter for j in range(7))
        self.nb7_penalty_costs = self.penalty_cost_nb7 * self.violation_nb7

        self.violation_nb8 = sum(self.solver.Value(self.nb8_violation[i, j]) for i in self.mitarbeiter for j in range(7, self.calc_time))
        self.nb8_penalty_costs = self.penalty_cost_nb8 * self.violation_nb8

        self.violation_nb9 = sum(self.solver.Value(self.nb9_violation[i, j, k]) for i in self.mitarbeiter for j in range(self.calc_time) for k in range(len(self.verfügbarkeit[i][j])))
        self.nb9_penalty_costs = (self.penalty_cost_nb9 / self.hour_devider) * self.violation_nb9

        self.violation_nb10 = sum(self.solver.Value(self.nb10_violation[i, j]) for i in self.mitarbeiter for j in range(self.calc_time))
        self.nb10_penalty_costs = self.penalty_cost_nb10 * self.violation_nb10

        self.violation_nb11 = sum(self.solver.Value(self.nb11_min_violation[i]) for i in self.mitarbeiter)
        self.nb11_penalty_costs = self.penalty_cost_nb11_min * self.violation_nb11

        self.violation_nb12 = sum(self.solver.Value(self.nb12_max_violation[i]) for i in self.mitarbeiter)
        self.nb12_penalty_costs = self.penalty_cost_nb12_max * self.violation_nb12



        # Kosten der einzelnen NBs ausgeben
        print('Kosten Einstellung von Mitarbeitern:', self.hiring_costs)
        print('Kosten Weiche NB1 (Mindestanzahl MA zu jeder Stunde an jedem Tag anwesend):', self.nb1_penalty_costs)
        print('Kosten Weiche NB2 (Max. Arbeitszeit pro Woche MA):', self.nb2_penalty_costs)
        print('Kosten Weiche NB3 (Min. Arbeitszeit pro Tag):', self.nb3_min_penalty_costs)
        print('Kosten Weiche NB4 (Max. Arbeitszeit pro Tag):', self.nb4_max_penalty_costs)
        print('Kosten Weiche NB5 (Unterschreitung der festen Mitarbeiter zu employment_level):', self.nb5_min_penalty_costs)
        print('Kosten Weiche NB6 (Überschreitung der festen Mitarbeiter zu employment_level):', self.nb6_max_penalty_costs)
        print('Kosten Weiche NB7 (Immer die gleiche Schicht in einer Woche):', self.nb7_penalty_costs)
        print('Kosten Weiche NB8 (Immer die gleiche Schicht zweite Woche):', self.nb8_penalty_costs)
        print('Kosten Weiche NB9 (Minimale Arbeitsstunden pro Block):', self.nb9_penalty_costs)
        print('Kosten Weiche NB10 (Max. Anzahl an Arbeitstagen in Folge):', self.nb10_penalty_costs)
        print('Kosten Weiche NB11 (Unterschreitung der gerechten Verteilung der Stunden):', self.nb11_penalty_costs)
        print('Kosten Weiche NB12 (Überschreitung der gerechten Verteilung der Stunden):', self.nb12_penalty_costs)
        print('Gesamtkosten:', self.solver.ObjectiveValue())



    def store_solved_data(self):
        """
        mitarbeiter_arbeitszeiten Attribut befülle
        """

        if self.status is None:
            print("Das Problem wurde noch nicht gelöst.")
            return

        """
        if self.status == cp_model.OPTIMAL or self.status == cp_model.FEASIBLE:
            for i in self.mitarbeiter:
                self.mitarbeiter_arbeitszeiten[i] = [
                    [self.solver.Value(self.x[i, j, k]) for k in range(len(self.verfügbarkeit[i][j]))] 
                    for j in range(self.calc_time)
                ]
            print(self.mitarbeiter_arbeitszeiten)
        """

        # 23.10.2023
        if self.status == cp_model.OPTIMAL or self.status == cp_model.FEASIBLE:
            for i in self.mitarbeiter:
                self.mitarbeiter_arbeitszeiten[i] = {
                    s: [
                        [
                            self.solver.Value(self.x[i, j, k, s]) if (i, j, k, s) in self.x else 0 
                            for k in range(len(self.verfügbarkeit[i][j]))
                        ]
                        for j in range(self.calc_time)
                    ]
                    for s in self.skills if s in self.mitarbeiter_s[i]
                }
            print(self.mitarbeiter_arbeitszeiten)



        if self.status == cp_model.OPTIMAL:
            print("Optimale Lösung gefunden.")
        elif self.status == cp_model.FEASIBLE:
            print("Mögliche Lösung gefunden.")
        elif self.status == cp_model.INFEASIBLE:
            print("Problem ist unlösbar.")
        elif self.status == cp_model.NOT_SOLVED:
            print("Solver hat das Problem nicht gelöst.")
        else:
            print("Unbekannter Status.")


    def output_result_excel(self):
        """
        Excel ausgabe
        """
        data = self.mitarbeiter_arbeitszeiten

        # Legen Sie die Füllungen fest
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Festlegen der Schriftgröße
        font = Font(size=10)
        header_font = Font(size=5)  # Schriftgröße für die Spaltentitel

        # Erstellen Sie ein Workbook
        wb = Workbook()
        ws = wb.active

        # Schreiben Sie die Überschriften
        headers = ["user_id"]
        for i in range(1, len(data[list(data.keys())[0]]) + 1):
            headers.extend(["T{}, {}:{}".format(i, j+8, k*15) for j in range(10) for k in range(4)])
            headers.append(' ')
        headers.append("Total Hours") 
        ws.append(headers)

        # Ändern der Schriftgröße der Spaltentitel
        for cell in ws[1]:
            cell.font = header_font

        # Schreiben Sie die Daten
        for idx, (ma, days) in enumerate(data.items(), start=2):
            row = [ma]
            for day in days:
                quarter_hours = [h for h in day]
                row.extend(quarter_hours)
                row.append(' ')  # Fügt eine leere Spalte nach jedem Tag hinzu
            ws.append(row)
            total_hours_col = len(headers)  # Wir verwenden die Anzahl der Überschriften, um die richtige Spalte zu erhalten
            ws.cell(row=idx, column=total_hours_col, value=f"=SUM(B{idx}:{get_column_letter(total_hours_col-1)}{idx})")


        # Farben auf Basis der Zellenwerte festlegen und Schriftgröße für den Rest des Dokuments
        for row in ws.iter_rows(min_row=2, values_only=False):
            for cell in row[1:]:
                if cell.value == 1:
                    cell.fill = green_fill
                elif cell.value == 0:
                    cell.fill = red_fill
                cell.font = font

        # Ändern der Spaltenbreite
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 3:
                adjusted_width = 3
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        wb.save("Einsatzplan.xlsx")


    def plot_costs_excel(self):
        """
        In dieser Methode werden die besten Kosten nach X Sekunden als Liniendiagramm in einem Excel gespeichert
        """

        # Workbook erstellen
        wb = Workbook()
        ws = wb.active
        ws.title = "Kosten"

        # Datei in das Arbeitsblatt einfügen
        data = [
            ['Zeit (Sekunden)', 'Beste Werte', 'Optimale Werte'],
        ]
        times = [i * self.best_test_time for i in range(1, len(self.best_values) + 1)]
        for i in range(len(times)):
            data.append([times[i], self.best_values[i], (self.hiring_costs - self.nb1_penalty_costs)]) # Penaltykosten von NB1 werden abgezogen, damit nur die minimale Einstellung der MA berechnet wird

        for row in data:
            ws.append(row)

        # Liniendiagramm erstellen
        chart = LineChart()
        chart.title = "Kosten über die Zeit"
        chart.style = 13
        chart.x_axis.title = 'Zeit (Sekunden)'
        chart.y_axis.title = 'Kosten'

        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=len(self.best_values)+1)
        chart.add_data(data, titles_from_data=True)
        times = Reference(ws, min_col=1, min_row=2, max_row=len(self.best_values)+1)
        chart.set_categories(times)
        ws.add_chart(chart, "E5")

        wb.save("Kostenanalyse.xlsx")


    def save_data_in_database(self):
        """ 
        Diese Methode speichert die berechneten Arbeitszeiten in der Datenbank 
        """
        with app.app_context():
            
            # Bestimme den maximalen Datumsbereich aller Benutzer
            all_start_dates = [self.user_availability[user_id][0][0] for user_id in self.mitarbeiter_arbeitszeiten]
            all_end_dates = [start_date + datetime.timedelta(days=max(len(days) for days in departments.values()) - 1)
                            for start_date, departments in zip(all_start_dates, self.mitarbeiter_arbeitszeiten.values())]

            # Wählen Sie das früheste Startdatum und das späteste Enddatum für den gesamten Löschvorgang
            print(all_start_dates)
            print(all_end_dates)
            
            global_start_date = min(all_start_dates)
            global_end_date = max(all_end_dates)
            print(global_start_date)
            print(global_end_date)

            # Lösche alle Einträge im globalen Datumsbereich
            Timetable.query.filter(
                Timetable.date >= global_start_date, 
                Timetable.date <= global_end_date
            ).delete()
            db.session.commit()

            for user_id, departments in self.mitarbeiter_arbeitszeiten.items(): # Durch mitarbeiter_arbeitszeiten durchitterieren
                print(f"Verarbeite Benutzer-ID: {user_id}")

                # Benutzer aus der Datenbank abrufen
                user = User.query.get(user_id)
                print(user)
                if not user:
                    print(f"Kein Benutzer gefunden mit ID: {user_id}")
                    continue

                for department, days in departments.items():
                    print(f"Verarbeite Abteilung: {department}")

                    for day_index, day in enumerate(days):
                        # Wir gehen davon aus, dass der erste Tag im 'self.user_availability' das Startdatum ist
                        date = self.user_availability[user_id][0][0] + datetime.timedelta(days=day_index)
                        weekday = date.strftime('%A') # Den aktuellen Wochentag finden

                        # Hier unterteilen wir den Tag in Schichten, basierend auf den Zeiten, zu denen der Mitarbeiter arbeitet
                        shifts = []
                        start_time_index = None

                        for time_index in range(len(day)):
                            if day[time_index] == 1 and start_time_index is None:
                                start_time_index = time_index
                            elif day[time_index] == 0 and start_time_index is not None:
                                shifts.append((start_time_index, time_index))
                                start_time_index = None

                        if start_time_index is not None:
                            shifts.append((start_time_index, len(day)))

                        print(f"Berechnete Schichten für Benutzer-ID {user_id}, Tag-Index {day_index}: {shifts}")

                        # Divisor bestimmen
                        divisor = 3600 / self.hour_devider

                        for shift_index, (start_time, end_time) in enumerate(shifts):
                            opening_time_in_units = int(self.laden_oeffnet[day_index].total_seconds() * self.hour_devider / 3600)
                            start_time += opening_time_in_units
                            end_time += opening_time_in_units

                            start_datetime = datetime.datetime.combine(date, datetime.datetime.min.time()) + timedelta(hours=start_time // self.hour_devider, minutes=(start_time % self.hour_devider) * 60 / self.hour_devider)
                            end_datetime = datetime.datetime.combine(date, datetime.datetime.min.time()) + timedelta(hours=end_time // self.hour_devider, minutes=(end_time % self.hour_devider) * 60 / self.hour_devider)

                            # Überprüfen Sie, ob das Enddatum auf den nächsten Tag überläuft
                            if end_datetime.time() < start_datetime.time():
                                end_datetime += timedelta(days=1)

                            # Neues Timetable-Objekt
                            if shift_index == 0:
                                new_entry = Timetable(
                                    id=None,
                                    email=user.email,
                                    first_name=user.first_name,
                                    last_name=user.last_name,
                                    company_name=user.company_name,
                                    department=department,
                                    date=date,
                                    weekday = weekday,
                                    start_time=start_datetime,
                                    end_time=end_datetime,
                                    start_time2=None,
                                    end_time2=None,
                                    start_time3=None,
                                    end_time3=None,
                                    created_by=self.current_user_id,
                                    changed_by=self.current_user_id,
                                    creation_timestamp=datetime.datetime.now()
                                )
                                db.session.add(new_entry)

                            elif shift_index == 1:
                                new_entry.start_time2 = datetime.datetime.combine(date, datetime.time(hour=int(start_time // self.hour_devider), minute=int((start_time % self.hour_devider) * 60 / self.hour_devider)))
                                new_entry.end_time2 = datetime.datetime.combine(date, datetime.time(hour=int(end_time // self.hour_devider), minute=int((end_time % self.hour_devider) * 60 / self.hour_devider)))

                            # new_entry der Datenbank hinzufügen
                            db.session.add(new_entry)

                # Änderungen in der Datenbank speichern
                db.session.commit()


    def save_data_in_database_testing(self):
        """ 
        Diese Methode speichert verwendete Daten in der Datenbank für das testing
        """
        new_entry = SolverAnalysis(
            id = None,
            usecase = self.solver_requirements["company_name"],
            self_current_user_id = self.current_user_id,
            self_user_availability = str(self.user_availability),
            self_opening_hours = str(self.opening_hours),
            self_laden_oeffnet = str(self.laden_oeffnet),
            self_laden_schliesst = str(self.laden_schliesst),
            self_binary_availability = str(self.binary_availability),
            self_company_shifts = self.company_shifts,
            self_weekly_hours = self.weekly_hours,
            self_employment_lvl = str(self.employment_lvl),
            self_time_req = str(self.time_req),
            self_user_employment = self.user_employment,
            self_solver_requirements = str(self.solver_requirements),
            self_week_timeframe = self.week_timeframe,
            self_hour_devider = self.hour_devider,
            self_mitarbeiter = str(self.mitarbeiter),
            self_verfügbarkeit = str(self.verfügbarkeit),
            self_kosten = str(self.kosten),
            self_max_zeit = str(self.max_zeit),
            self_min_zeit = str(self.min_zeit),
            self_max_time_week = self.max_time_week,
            self_calc_time = self.calc_time,
            self_employment_lvl_exact = str(self.employment_lvl_exact),
            self_employment = str(self.employment),
            self_verteilbare_stunden = self.verteilbare_stunden,
            self_gesamtstunden_verfügbarkeit = str(self.gesamtstunden_verfügbarkeit),
            self_min_anwesend = str(self.min_anwesend),
            self_gerechte_verteilung = str(self.gerechte_verteilung),
            self_fair_distribution = self.fair_distribution,
            solving_time = self.solving_time_seconds,
            lp_iteration = None,
            violation_nb1 = self.violation_nb1,
            violation_nb2 = self.violation_nb2,
            violation_nb3 = self.violation_nb3,
            violation_nb4 = self.violation_nb4,
            violation_nb5 = self.violation_nb5,
            violation_nb6 = self.violation_nb6,
            violation_nb7 = self.violation_nb7,
            violation_nb8 = self.violation_nb8,
            violation_nb9 = self.violation_nb9,
            violation_nb10 = self.violation_nb10,
            violation_nb11 = self.violation_nb11,
            violation_nb12 = self.violation_nb12,
            violation_nb13 = None,
            violation_nb14 = None,
            violation_nb15 = None,
            violation_nb16 = None,
            violation_nb17 = None,
            violation_nb18 = None,
            violation_nb19 = None,
            violation_nb20 = None,
            penalty_cost_nb1 = self.penalty_cost_nb1,
            penalty_cost_nb2 = self.penalty_cost_nb2,
            penalty_cost_nb3 = self.penalty_cost_nb3_min,
            penalty_cost_nb4 = self.penalty_cost_nb4_max,
            penalty_cost_nb5 = self.penalty_cost_nb5_min,
            penalty_cost_nb6 = self.penalty_cost_nb6_max,
            penalty_cost_nb7 = self.penalty_cost_nb7,
            penalty_cost_nb8 = self.penalty_cost_nb8,
            penalty_cost_nb9 = self.penalty_cost_nb9,
            penalty_cost_nb10 = self.penalty_cost_nb10,
            penalty_cost_nb11 = self.penalty_cost_nb11_min,
            penalty_cost_nb12 = self.penalty_cost_nb12_max,
            penalty_cost_nb13 = None,
            penalty_cost_nb14 = None,
            penalty_cost_nb15 = None,
            penalty_cost_nb16 = None,
            penalty_cost_nb17 = None,
            penalty_cost_nb18 = None,
            penalty_cost_nb19 = None,
            penalty_cost_nb20 = None,
            possible_solution = None,
            gap_016 = None,
            gap_05 = None,
            gap_1 = None,
            gap_2 = None,
            gap_3 = None,
            gap_4 = None,
            gap_5 = None,
            gap_10 = None,
            gap_20 = None,
            gap_30 = None,
            memory = None
        )

        # new_entry der Datenbank hinzufügen
        db.session.add(new_entry)

        # Änderungen in der Datenbank speichern
        db.session.commit()
        print("Daten erfolgreich in die Datenbank gespeichert.")