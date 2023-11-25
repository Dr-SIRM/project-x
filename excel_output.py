import io

from datetime import datetime, timedelta, time
from models import Timetable
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from app import app, db, timedelta, get_database_uri
from sqlalchemy import text, case
from models import User, Availability, TimeReq, Company, OpeningHours, Timetable, \
    TemplateAvailability, TemplateTimeRequirement, RegistrationToken, PasswordReset, \
    SolverRequirement, SolverAnalysis
from flask_jwt_extended import get_jwt
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine, not_


def get_session(uri):
    engine = create_engine(uri)
    Session = sessionmaker(bind=engine)
    return Session()

def time_to_timedelta(t):
    return timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)


def create_excel_output(current_user_email, start, end):
    """
    In dieser Funktion werden relevante gesolvte Daten aus der Datenbank gezogen und eine Excelausgabe daraus generiert.
    """

    # Achtung, start_date muss immer ein Montag sein!
    start_date = str(start)
    end_date = str(end)


    # Neue Session starten
    jwt_data = get_jwt()
    session_company = jwt_data.get("company_name").lower().replace(' ', '_')
    session = get_session(get_database_uri('', session_company))


    """
    COMPANY INFOS ZIEHEN ---------------------------------------------------------------
    """
    # Hole den company_name des aktuellen Benutzers
    user = session.query(User).filter_by(email=current_user_email).first()
    company_name = user.company_name

    # Liste von Farben für die Departments
    available_colors = [
        "B0C4DE",  # Sanftes Blau
        "98FB98",  # Zartes Grün
        "FFFF99",  # Helles Gelb
        "FF9999",  # Sanftes Rot
        "87CEEB",  # Himmelblau
        "E6E6FA",  # Lavendel
        "FFEFD5",  # Pfirsich
        "98FF98",  # Mintgrün
        "FFB6C1",  # Babyrosa
        "FFFDD0",  # Creme
    ]


    # Die entsprechende Company-Instanz abfragen
    company = session.query(Company).filter_by(company_name=company_name).first()

    # Liste für die Abteilungen
    departments = []

    # Zugriff auf das erste Department
    first_department = getattr(company, 'department')
    if first_department:
        departments.append(first_department)

    # Die restlichen Departments durchiterieren 
    for i in range(2, 11):
        department_value = getattr(company, f'department{i}')
        if department_value:
            departments.append(department_value)

    # Zuweisen von Farben zu Abteilungen
    department_colors = {}
    for i, department in enumerate(departments):
        department_initials = department[:2].upper()  # Nehmen Sie die ersten zwei Buchstaben in Großbuchstaben
        department_colors[department_initials] = available_colors[i % len(available_colors)]
    print("department_colors: ", department_colors)

    # ----------------------------------------------------------------------------------

    # Definieren Sie eine case Bedingung, um die Tage der Woche in der richtigen Reihenfolge zu sortieren
    weekday_order = case(
        (OpeningHours.weekday == "Montag", 1),
        (OpeningHours.weekday == "Dienstag", 2),
        (OpeningHours.weekday == "Mittwoch", 3),
        (OpeningHours.weekday == "Donnerstag", 4),
        (OpeningHours.weekday == "Freitag", 5),
        (OpeningHours.weekday == "Samstag", 6),
        (OpeningHours.weekday == "Sonntag", 7),
        else_=100 
    )

    # Öffnungszeiten basierend auf dem company_name abrufen und nach Wochentagen sortieren
    opening = session.query(OpeningHours).filter_by(company_name=company_name).order_by(weekday_order).all()

    # Konvertieren Sie die Zeiten in time_delta und speichern Sie sie zusammen mit dem Wochentag in einer Liste
    times = [
        (record.weekday, 
         time_to_timedelta(record.start_time), 
         time_to_timedelta(record.end_time), 
         time_to_timedelta(record.end_time2) if record.end_time2 is not None else None) 
        for record in opening
    ]
    
    # Differenz zwischen den Daten in Tagen berechnen
    date_diff = datetime.strptime(end_date, "%Y-%m-%d") - datetime.strptime(start_date, "%Y-%m-%d")
    weeks_diff = date_diff.days // 7

    # Erweitern der times Liste für jede zusätzliche Woche
    extended_times = times[:]
    for i in range(1, weeks_diff + 1):
        for day in times:
            new_day = (
                day[0],  # Wochentag bleibt gleich
                day[1],  # start_time bleibt gleich
                day[2],  # end_time bleibt gleich
                day[3]   # end_time2 bleibt gleich, falls vorhanden
            )
            extended_times.append(new_day)

    # Wieder in times umwadeln
    times = extended_times

    # ----------------------------------------------------------------------------------

    # Abfrage, um hour_divider abzurufen
    solver_requirement = session.query(SolverRequirement).filter_by(company_name=company_name).first()
    hour_divider = solver_requirement.hour_divider if solver_requirement else None


    """
    USER INFOS ZIEHEN ------------------------------------------------------------------
    """

    # Abfrage, um die User-Informationen abzurufen
    company_users = session.query(User).filter_by(company_name=company_name).with_entities(
        User.id, User.first_name, User.last_name, User.employment, 
        User.email, User.employment_level
        ).filter(not_(User.access_level == "Super_Admin")
    ).all()

    # ----------------------------------------------------------------------------------

    # Abfrage, um die Timetable-Informationen abzurufen
    data_timetable = (
        session.query(Timetable).filter_by(company_name=company_name)
        .filter(Timetable.date.between(start_date, end_date))
        .with_entities(
            Timetable.email,
            Timetable.department,
            Timetable.date,
            Timetable.start_time,
            Timetable.end_time,
            Timetable.start_time2,
            Timetable.end_time2,
        )
        .all()
    )

    # Konvertieren Sie die Zeiten in timedelta und speichern Sie sie in einer Liste
    data_timetable = [
        (
            email,
            department, 
            date, 
            time_to_timedelta(start_time), 
            time_to_timedelta(end_time), 
            time_to_timedelta(start_time2) if start_time2 is not None else None, 
            time_to_timedelta(end_time2) if end_time2 is not None else None
        ) 
        for email, department, date, start_time, end_time, start_time2, end_time2 in data_timetable
    ]


    # Excel erstellen und befüllen ----------------------------------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active

    # Erstelle einen Border-Style --> Rahmenlinien definieren
    border_style = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Company Name in Zelle B2 eintragen mit Schriftgröße 24
    ws['B2'] = f"Einsatzplan {company_name}"  # company_name ist der Name der Firma, der aus der Datenbank abgerufen wurde
    ws['B2'].font = Font(size=24)  # Schriftgröße auf 24

    # Spaltentitel setzen
    titles = [
        (2, "ID"),
        (3, "Vorname"),
        (4, "Name"),
        (5, "Anstellung"),
        (6, "Anstellungsgrad")
    ]
    
    fill_color = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx, title in titles:
        cell = ws.cell(row=5, column=col_idx, value=title)
        cell.alignment = Alignment(horizontal='left')  # Setze den Text des Titels linksbündig
        cell.fill = fill_color  # Setze die Füllfarbe der Zelle
        cell.border = border_style  # Rahmenlinien für die Titel setzen

        
        # Spaltenbreiten einstellen
        if title == "ID":
            ws.column_dimensions[chr(64 + col_idx)].width = 5
        if title == "Vorname":
            ws.column_dimensions[chr(64 + col_idx)].width = 12
        if title == "Name":
            ws.column_dimensions[chr(64 + col_idx)].width = 12
        if title == "Anstellung":
            ws.column_dimensions[chr(64 + col_idx)].width = 12
        if title == "Anstellungsgrad":
            ws.column_dimensions[chr(64 + col_idx)].width = 15
    

    # Startzeile für Daten
    first_data_row = 6
    row = first_data_row

    email_row_dict = {}
    # User-Informationen eintragen
    for user in company_users:
        user_id, first_name, last_name, employment, email, employment_level = user
        email_row_dict[email] = row
        data = [
            (2, user_id),
            (3, first_name),
            (4, last_name),
            (5, employment),
            (6, employment_level)
        ]
        for col_idx, value in data:
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left')  # Setze den Text der Datenzelle linksbündig
            cell.border = border_style  # Rahmenlinien für die User-Informationen setzen
        row += 1

    last_data_row = row - 1  # Da row nach dem letzten Eintrag erhöht wurde, reduzieren wir es um 1.


    # Funktion zur Generierung der Stunden
    def generate_hours(start_time, end_time, hour_divider):
        current_time = start_time
        times = []
        delta = timedelta(minutes=60//hour_divider)
        while current_time < end_time:
            times.append(current_time.strftime('%H:%M'))
            current_time += delta
        return times
    
    # Stunden für jeden Tag hinzufügen
    col_index = 7  # Beginne in Spalte G

    for time_info in times:
        weekday, start_time_obj, end_time_obj, end_time2_obj = time_info
        start_time = (datetime.min + start_time_obj).time()
        if end_time2_obj is None:
            end_time = (datetime.min + end_time_obj).time()
        else:
            end_time = (datetime.min + end_time2_obj).time()

        # hours = generate_hours(datetime.combine(datetime.today(), start_time), datetime.combine(datetime.today(), end_time), hour_divider)
        # Erstellen eines datetime-Objekts für start_time und end_time
        start_datetime = datetime.combine(datetime.today(), start_time)
        end_datetime = datetime.combine(datetime.today(), end_time)

        # Wenn end_time kleiner als start_time ist, fügen Sie einen Tag zu end_datetime hinzu
        if end_datetime < start_datetime:
            end_datetime += timedelta(days=1)

        hours = generate_hours(start_datetime, end_datetime, hour_divider)

        row_index = 5
        
        if hours:
            day_cell = ws.cell(row=4, column=col_index, value=f"{weekday} {start_date}")
            if len(hours) > 1:
                ws.merge_cells(start_row=4, start_column=col_index, end_row=4, end_column=col_index + len(hours) - 1)
                
            for merge_col_index in range(col_index, col_index + len(hours)):
                cell = ws.cell(row=4, column=merge_col_index)
                cell.border = border_style
                
            for hour in hours:
                cell = ws.cell(row=row_index, column=col_index, value=hour)
                cell.font = Font(size=6)
                cell.border = border_style
                ws.column_dimensions[get_column_letter(col_index)].width = 3.2
                col_index += 1
            
            for email, department, date, start_time, end_time, start_time2, end_time2 in data_timetable:
                if date.strftime("%Y-%m-%d") == start_date:
                    user_row = email_row_dict.get(email)
                    if user_row:
                        # Verarbeitung der ersten Schicht
                        start_time = (datetime.min + start_time).time()
                        end_time = (datetime.min + end_time).time()
                        
                        start_col = col_index - len(hours) + (datetime.combine(datetime.today(), start_time) - datetime.combine(datetime.today(), (datetime.min + start_time_obj).time())).seconds // (60 * 60 // hour_divider)
                        end_col = start_col + (datetime.combine(datetime.today(), end_time) - datetime.combine(datetime.today(), start_time)).seconds // (60 * 60 // hour_divider) - 1

                        # Färben der Zellen, Einfügen der Abteilungsinitialen und Formatierung
                        for col in range(start_col, end_col + 1):
                            cell = ws.cell(row=user_row, column=col)
                            department_initials = department[:2].upper()  # Erste zwei Buchstaben der Abteilung
                            cell.value = department_initials  # Setze die Abteilungsinitialen in die Zelle
                            cell.alignment = Alignment(horizontal='center', vertical='center')  # Zentriere den Text
                            cell.font = Font(size=8, bold=True)  # Setze die Schriftgröße und Schriftart

                            cell_color = department_colors.get(department_initials, "00FF00")  # Standardfarbe ist Grün
                            cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")


                        # Verarbeitung der zweiten Schicht
                        if start_time2 and end_time2:
                            start_time2 = (datetime.min + start_time2).time()
                            end_time2 = (datetime.min + end_time2).time()

                            start_col2 = col_index - len(hours) + (datetime.combine(datetime.today(), start_time2) - datetime.combine(datetime.today(), (datetime.min + start_time_obj).time())).seconds // (60 * 60 // hour_divider)
                            end_col2 = start_col2 + (datetime.combine(datetime.today(), end_time2) - datetime.combine(datetime.today(), start_time2)).seconds // (60 * 60 // hour_divider) - 1

                            # Färben der Zellen, Einfügen der Abteilungsinitialen und Formatierung für die zweite Schicht
                            for col in range(start_col2, end_col2 + 1):
                                cell = ws.cell(row=user_row, column=col)
                                department_initials = department[:2].upper()  # Erste zwei Buchstaben der Abteilung
                                cell.value = department_initials  # Setze die Abteilungsinitialen in die Zelle
                                cell.alignment = Alignment(horizontal='center', vertical='center')  # Zentriere den Text
                                cell.font = Font(size=8, bold=True)  # Setze die Schriftgröße und Schriftart

                                cell_color = department_colors.get(department_initials, "00FF00")  # Standardfarbe ist Grün
                                cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")

        start_date = (datetime.strptime(start_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")


    # Rahmen für alle Zellen
    for row in range(first_data_row, last_data_row + 1):
        for col in range(7, col_index):
            cell = ws.cell(row=row, column=col)
            cell.border = border_style  # Setzen Sie hier den vorher definierten border_style.



    # Legende für die Abteilungen hinzufügen
    legend_row = last_data_row + 2  # Beginne zwei Zeilen unter der letzten Datenzeile
    ws.cell(row=legend_row, column=2, value="Legende Abteilungen:").font = Font(size=11, bold=True)

    for i, department in enumerate(departments):
        legend_row += 1  # Für jede Abteilung eine neue Zeile
        department_initials = department[:2].upper()

        # Füge die Abteilungsinitialen und die zugehörige Farbe in die Zelle ein
        initial_cell = ws.cell(row=legend_row, column=2, value=department_initials)
        initial_cell.font = Font(size=8, bold=True)
        initial_cell.alignment = Alignment(horizontal='center')
        cell_color = department_colors.get(department_initials, "00FF00")  # Standardfarbe ist Grün
        initial_cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")

        # Füge den vollständigen Namen der Abteilung in die nächste Zelle rechts ein
        ws.cell(row=legend_row, column=3, value=department).font = Font(size=10)


    # In-memory bytes stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)  # Gehe zum Start des Streams

    # Session beenden
    session.close()

    return output



